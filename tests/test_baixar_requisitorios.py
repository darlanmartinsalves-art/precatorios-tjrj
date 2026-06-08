"""Testes do script baixar_requisitorios."""
import pytest
from baixar_requisitorios import split_cnj


def test_split_cnj_padrao_valido():
    assert split_cnj("0156129-30.2020.8.19.0001") == ("0156129-30.2020", "0001")


def test_split_cnj_outro_processo():
    assert split_cnj("0230764-50.2018.8.19.0001") == ("0230764-50.2018", "0001")


def test_split_cnj_vara_diferente_0072():
    assert split_cnj("0000224-76.2021.8.19.0072") == ("0000224-76.2021", "0072")


def test_split_cnj_formato_invalido_levanta():
    with pytest.raises(ValueError):
        split_cnj("formato-invalido")


def test_split_cnj_vazio_levanta():
    with pytest.raises(ValueError):
        split_cnj("")


from baixar_requisitorios import sanitizar_nome


def test_sanitizar_remove_caracteres_invalidos_windows():
    assert sanitizar_nome('foo/bar\\baz:qux*?"<>|') == "foo_bar_baz_qux______"


def test_sanitizar_colapsa_espacos_multiplos():
    assert sanitizar_nome("ABC    DEF") == "ABC DEF"


def test_sanitizar_strip_pontas():
    assert sanitizar_nome("   FOO   ") == "FOO"


def test_sanitizar_limita_100_chars():
    longo = "X" * 200
    resultado = sanitizar_nome(longo)
    assert len(resultado) == 100


def test_sanitizar_nome_real_beneficiario():
    nome = "TECHNE ENGENHARIA E SISTEMAS LTDA"
    assert sanitizar_nome(nome) == "TECHNE ENGENHARIA E SISTEMAS LTDA"


from baixar_requisitorios import extrair_beneficiario


def test_extrair_beneficiario_do_pdf_modelo(pdf_modelo_bytes):
    resultado = extrair_beneficiario(pdf_modelo_bytes)
    assert resultado == "TECHNE ENGENHARIA E SISTEMAS LTDA"


def test_extrair_beneficiario_pdf_vazio_retorna_none():
    # PDF mínimo válido sem conteúdo de beneficiário
    pdf_minimo = b"%PDF-1.4\n%%EOF"
    assert extrair_beneficiario(pdf_minimo) is None


def test_extrair_beneficiario_dados_nao_pdf_retorna_none():
    assert extrair_beneficiario(b"isso nao eh um pdf") is None


from baixar_requisitorios import gerar_nome_arquivo
from pathlib import Path


def test_gerar_nome_arquivo_sem_colisao(tmp_path):
    resultado = gerar_nome_arquivo("2025.09451-0", "TECHNE ENGENHARIA", tmp_path)
    assert resultado == tmp_path / "2025.09451-0 - TECHNE ENGENHARIA.pdf"


def test_gerar_nome_arquivo_com_colisao_simples(tmp_path):
    (tmp_path / "2025.09451-0 - TECHNE.pdf").write_bytes(b"")
    resultado = gerar_nome_arquivo("2025.09451-0", "TECHNE", tmp_path)
    assert resultado == tmp_path / "2025.09451-0 - TECHNE (2).pdf"


def test_gerar_nome_arquivo_com_colisao_dupla(tmp_path):
    (tmp_path / "2025.09451-0 - TECHNE.pdf").write_bytes(b"")
    (tmp_path / "2025.09451-0 - TECHNE (2).pdf").write_bytes(b"")
    resultado = gerar_nome_arquivo("2025.09451-0", "TECHNE", tmp_path)
    assert resultado == tmp_path / "2025.09451-0 - TECHNE (3).pdf"


def test_gerar_nome_arquivo_sanitiza_beneficiario(tmp_path):
    resultado = gerar_nome_arquivo("2025.09451-0", "TECHNE/INC", tmp_path)
    assert resultado == tmp_path / "2025.09451-0 - TECHNE_INC.pdf"


from baixar_requisitorios import filtrar_precatorios


def test_filtrar_aplica_saldo_minimo(xlsx_com_saldos):
    resultado = filtrar_precatorios(xlsx_com_saldos, saldo_minimo=200000)
    numeros = sorted(p for p, _ in resultado)
    # 2025.00001 (100k) e 2025.00005 (199.999,99) ficam de fora; 2025.00006 (None) também
    assert numeros == ["2025.00002-1", "2025.00003-2", "2025.00004-3"]


def test_filtrar_retorna_precatorio_e_processo(xlsx_com_saldos):
    resultado = filtrar_precatorios(xlsx_com_saldos, saldo_minimo=200000)
    assert ("2025.00002-1", "0000002-00.2020.8.19.0001") in resultado
    assert ("2025.00003-2", "0000003-00.2020.8.19.0001") in resultado
    assert ("2025.00004-3", "0000003-00.2020.8.19.0001") in resultado


def test_filtrar_ignora_saldos_nao_numericos(xlsx_com_saldos):
    resultado = filtrar_precatorios(xlsx_com_saldos, saldo_minimo=0)
    # 2025.00006-5 tem saldo None, deve ser ignorado mesmo com mínimo 0
    numeros = [p for p, _ in resultado]
    assert "2025.00006-5" not in numeros


from baixar_requisitorios import carregar_checkpoint_dl, salvar_checkpoint_dl


def test_checkpoint_dl_inexistente_retorna_vazio(tmp_path):
    assert carregar_checkpoint_dl(tmp_path / "x.json") == {}


def test_checkpoint_dl_roundtrip(tmp_path):
    caminho = tmp_path / "state.json"
    dados = {"0000001-00.2020.8.19.0001": {"status": "ok", "arquivos": ["a.pdf"]}}
    salvar_checkpoint_dl(caminho, dados)
    assert carregar_checkpoint_dl(caminho) == dados


def test_checkpoint_dl_e_atomico_nao_deixa_tmp(tmp_path):
    caminho = tmp_path / "state.json"
    salvar_checkpoint_dl(caminho, {"a": "b"})
    assert not (tmp_path / "state.json.tmp").exists()
    assert caminho.exists()


from baixar_requisitorios import extrair_beneficiario_completo


def test_extrair_beneficiario_completo_do_modelo(pdf_modelo_bytes):
    resultado = extrair_beneficiario_completo(pdf_modelo_bytes)
    assert resultado == {
        "nome": "TECHNE ENGENHARIA E SISTEMAS LTDA",
        "doc": "50737766000121",
        "tipo_doc": "CNPJ",
    }


def test_extrair_beneficiario_completo_pdf_invalido_retorna_none():
    assert extrair_beneficiario_completo(b"nao eh pdf") is None


def test_extrair_beneficiario_completo_pdf_vazio_retorna_none():
    assert extrair_beneficiario_completo(b"%PDF-1.4\n%%EOF") is None


def test_extrair_beneficiario_legado_ainda_retorna_nome(pdf_modelo_bytes):
    # Função legada continua funcionando
    assert extrair_beneficiario(pdf_modelo_bytes) == "TECHNE ENGENHARIA E SISTEMAS LTDA"


from baixar_requisitorios import extrair_advogado


def test_extrair_advogado_do_modelo(pdf_modelo_bytes):
    resultado = extrair_advogado(pdf_modelo_bytes)
    assert resultado == {
        "nome": "ARTUR GARRASTAZU GOMES FERREIRA",
        "cpf": "33394784068",
        "oab": "RJ185918",
    }


def test_extrair_advogado_pdf_invalido_retorna_none():
    assert extrair_advogado(b"nao eh pdf") is None


def test_extrair_advogado_pdf_vazio_retorna_none():
    assert extrair_advogado(b"%PDF-1.4\n%%EOF") is None


from baixar_requisitorios import formatar_doc


def test_formatar_doc_cpf():
    assert formatar_doc("33394784068", "CPF") == "333.947.840-68"


def test_formatar_doc_cnpj():
    assert formatar_doc("50737766000121", "CNPJ") == "50.737.766/0001-21"


def test_formatar_doc_cpf_ja_formatado_retorna_formatado():
    assert formatar_doc("333.947.840-68", "CPF") == "333.947.840-68"


def test_formatar_doc_tipo_invalido_retorna_doc_cru():
    assert formatar_doc("12345", "OUTRO") == "12345"


def test_formatar_doc_doc_com_tamanho_errado_retorna_doc_cru():
    # CPF com 10 dígitos não casa
    assert formatar_doc("1234567890", "CPF") == "1234567890"


from baixar_requisitorios import atualizar_planilha


def test_atualizar_planilha_grava_colunas_N_a_R(xlsx_para_atualizar, tmp_path):
    saida = tmp_path / "saida.xlsx"
    dados = {
        "2025.00001-0": {
            "beneficiario_nome": "TECHNE ENGENHARIA",
            "beneficiario_doc": "50.737.766/0001-21",
            "advogado_nome": "ARTUR GARRASTAZU",
            "advogado_cpf": "333.947.840-68",
            "advogado_oab": "RJ185918",
        }
    }
    atualizar_planilha(xlsx_para_atualizar, saida, dados)

    from openpyxl import load_workbook
    wb = load_workbook(saida)
    ws = wb.active
    assert ws.cell(row=2, column=14).value == "TECHNE ENGENHARIA"    # N
    assert ws.cell(row=2, column=15).value == "50.737.766/0001-21"   # O
    assert ws.cell(row=2, column=16).value == "ARTUR GARRASTAZU"     # P
    assert ws.cell(row=2, column=17).value == "333.947.840-68"       # Q
    assert ws.cell(row=2, column=18).value == "RJ185918"             # R


def test_atualizar_planilha_preserva_entrada(xlsx_para_atualizar, tmp_path):
    import hashlib
    saida = tmp_path / "saida.xlsx"
    hash_antes = hashlib.md5(xlsx_para_atualizar.read_bytes()).hexdigest()
    atualizar_planilha(xlsx_para_atualizar, saida, {"2025.00001-0": {
        "beneficiario_nome": "X", "beneficiario_doc": "Y",
        "advogado_nome": "Z", "advogado_cpf": "W", "advogado_oab": "V",
    }})
    hash_depois = hashlib.md5(xlsx_para_atualizar.read_bytes()).hexdigest()
    assert hash_antes == hash_depois


def test_atualizar_planilha_preserva_colunas_existentes(xlsx_para_atualizar, tmp_path):
    saida = tmp_path / "saida.xlsx"
    atualizar_planilha(xlsx_para_atualizar, saida, {"2025.00001-0": {
        "beneficiario_nome": "X", "beneficiario_doc": "Y",
        "advogado_nome": "Z", "advogado_cpf": "W", "advogado_oab": "V",
    }})
    from openpyxl import load_workbook
    wb = load_workbook(saida)
    ws = wb.active
    # Coluna I=Cedente, J=CPF, K=Celular existentes intactos
    assert ws.cell(row=2, column=9).value == "ALGUEM"
    assert ws.cell(row=2, column=10).value == "111.111.111-11"
    assert ws.cell(row=2, column=11).value == "(21) 11111-1111"


def test_atualizar_planilha_grava_cabecalhos_novos(xlsx_para_atualizar, tmp_path):
    saida = tmp_path / "saida.xlsx"
    atualizar_planilha(xlsx_para_atualizar, saida, {})
    from openpyxl import load_workbook
    wb = load_workbook(saida)
    ws = wb.active
    assert ws.cell(row=1, column=14).value == "Beneficiário Nome"
    assert ws.cell(row=1, column=15).value == "Beneficiário Doc"
    assert ws.cell(row=1, column=16).value == "Advogado Nome"
    assert ws.cell(row=1, column=17).value == "Advogado CPF"
    assert ws.cell(row=1, column=18).value == "Advogado OAB"


def test_atualizar_planilha_ignora_precatorio_inexistente(xlsx_para_atualizar, tmp_path):
    saida = tmp_path / "saida.xlsx"
    atualizar_planilha(xlsx_para_atualizar, saida, {"2099.99999-9": {
        "beneficiario_nome": "X", "beneficiario_doc": "Y",
        "advogado_nome": "Z", "advogado_cpf": "W", "advogado_oab": "V",
    }})
    from openpyxl import load_workbook
    wb = load_workbook(saida)
    ws = wb.active
    for row in range(2, 5):
        assert ws.cell(row=row, column=14).value is None


import asyncio
from baixar_requisitorios import _erro_iframe_ausente, LoginExpiradoError


class _FakeLocator:
    def __init__(self, count):
        self._count = count

    async def count(self):
        return self._count


class _FakePage:
    """Page mínima: locator(...).count() devolve quantos campos de senha existem.
    `url` permite testar a detecção de sessão-morta por redirect pra fora do portal."""
    def __init__(self, campos_senha,
                 url="https://www3.tjrj.jus.br/consultaprocessual/#/consultaportal"):
        self._campos_senha = campos_senha
        self.url = url

    def locator(self, selector):
        return _FakeLocator(self._campos_senha)


def test_erro_iframe_com_login_expirado_retorna_login_expirado():
    # Campo de senha presente => sessão caiu (portal voltou pra tela de login)
    page = _FakePage(campos_senha=1)
    erro = asyncio.run(_erro_iframe_ausente(page, "0837412-55.2022.8.19.0001"))
    assert isinstance(erro, LoginExpiradoError)


def test_erro_iframe_sem_login_expirado_retorna_runtimeerror():
    # Sem campo de senha E ainda no portal => falha genuína de navegação daquele processo
    page = _FakePage(campos_senha=0)
    erro = asyncio.run(_erro_iframe_ausente(page, "0837412-55.2022.8.19.0001"))
    assert isinstance(erro, RuntimeError)
    assert not isinstance(erro, LoginExpiradoError)


def test_erro_iframe_redirect_fora_do_portal_retorna_login_expirado():
    # Sessão expira -> portal redireciona pro site público (sem campo de senha).
    # Mesmo sem senha, sair do portal = sessão morta => LoginExpiradoError (parada limpa).
    page = _FakePage(campos_senha=0, url="https://www.tjrj.jus.br/")
    erro = asyncio.run(_erro_iframe_ausente(page, "0837412-55.2022.8.19.0001"))
    assert isinstance(erro, LoginExpiradoError)


from baixar_requisitorios import _ordenar_candidatos


def test_ordenar_candidatos_inclui_filhos_em_ordem():
    # Cada match traz [idx_proprio, ...idx_filhos]. O documento real pode ser o filho.
    matches = [{"ids": [5, 6]}, {"ids": [9, 10]}]
    assert _ordenar_candidatos(matches) == [5, 6, 9, 10]


def test_ordenar_candidatos_remove_duplicados_preservando_ordem():
    # Filho que também aparece como match de outro padrão não pode duplicar
    matches = [{"ids": [5, 6]}, {"ids": [6, 7]}]
    assert _ordenar_candidatos(matches) == [5, 6, 7]


def test_ordenar_candidatos_vazio():
    assert _ordenar_candidatos([]) == []


from baixar_requisitorios import _parece_escaneado


def test_parece_escaneado_texto_vazio():
    assert _parece_escaneado("") is True


def test_parece_escaneado_pdf_imagem_so_quebras():
    # PDF escaneado: pypdf devolve só quebras de linha
    assert _parece_escaneado("\n\n\n") is True


def test_parece_escaneado_capa_juntada_nao_eh_scan():
    # Capa de juntada tem bastante texto — NÃO é scan
    texto = ("Estado do Rio de Janeiro Poder Judiciário Tribunal de Justiça "
             "Fase: Juntada Tipo de Documento Petição Data da Juntada 08/04/2025")
    assert _parece_escaneado(texto) is False


def test_parece_escaneado_requisitorio_real_nao_eh_scan():
    texto = "OFÍCIO REQUISITÓRIO DE PAGAMENTO III - BENEFICIÁRIO Nome: FULANO"
    assert _parece_escaneado(texto) is False


import seletores


def test_seletor_prolongar_sessao_escopado_ao_modal():
    # Regressão: 'div.rodape-confirma' sozinho casa ~13 elementos na página e o
    # clique acerta o botão errado -> sessão não é prolongada -> cai no meio do run.
    # O botão PRECISA ser escopado ao modal de sessão inativa.
    assert seletores.MODAL_SESSAO_INATIVA in seletores.MODAL_SESSAO_PROLONGAR


from baixar_requisitorios import _eh_segunda_instancia


def test_eh_segunda_instancia_sufixo_0000():
    # Órgão de origem .0000 = 2º grau -> sistema legado ejud
    assert _eh_segunda_instancia("0090251-30.2021.8.19.0000") is True


def test_eh_segunda_instancia_primeira_instancia():
    assert _eh_segunda_instancia("0156129-30.2020.8.19.0001") is False


def test_eh_segunda_instancia_vazio():
    assert _eh_segunda_instancia("") is False


from baixar_requisitorios import _eh_falha_navegacao


def test_eh_falha_navegacao_erros():
    # Falhas que em sequência indicam navegador/sessão quebrado
    assert _eh_falha_navegacao("erro_navegacao") is True
    assert _eh_falha_navegacao("processo_nao_encontrado") is True


def test_eh_falha_navegacao_resultados_legitimos():
    # Resultados legítimos NÃO contam pro circuit breaker
    assert _eh_falha_navegacao("ok") is False
    assert _eh_falha_navegacao("sem_requisitorio") is False
    assert _eh_falha_navegacao("legado_2inst") is False
    assert _eh_falha_navegacao("processo_pje") is False


from baixar_requisitorios import _parece_logado_portal


def test_parece_logado_portal_url_autenticada():
    assert _parece_logado_portal(
        "https://www3.tjrj.jus.br/portalservicos/#/consproc/consultaportal") is True


def test_parece_logado_portal_site_publico_eh_falso():
    # Sem sessão, o portal redireciona pro site público — NÃO é estar logado
    assert _parece_logado_portal("https://www.tjrj.jus.br/") is False


def test_parece_logado_portal_tela_login_eh_falso():
    assert _parece_logado_portal("https://www3.tjrj.jus.br/portalservicos/login") is False


def test_parece_logado_portal_vazio():
    assert _parece_logado_portal("") is False


from baixar_requisitorios import extrair_numero_ofreq


def test_extrair_numero_ofreq_do_modelo(pdf_modelo_bytes):
    assert extrair_numero_ofreq(pdf_modelo_bytes) == "2025.14235"


def test_extrair_numero_ofreq_pdf_invalido_retorna_none():
    assert extrair_numero_ofreq(b"nao eh pdf") is None


from baixar_requisitorios import extrair_vinculo_ofreq_precatorio


def test_extrair_vinculo_do_modelo(pdf_vinculo_bytes):
    assert extrair_vinculo_ofreq_precatorio(pdf_vinculo_bytes) == ("2025.06478", "2025.06209-0")


def test_extrair_vinculo_pdf_requisitorio_retorna_none(pdf_modelo_bytes):
    # O requisitório NAO eh documento de vinculo (nao tem "gerou o precatorio")
    assert extrair_vinculo_ofreq_precatorio(pdf_modelo_bytes) is None


def test_extrair_vinculo_pdf_invalido_retorna_none():
    assert extrair_vinculo_ofreq_precatorio(b"nao eh pdf") is None


from baixar_requisitorios import (
    eh_documento_requisitorio,
    eh_documento_vinculo,
    _extrair_texto_pdf,
)


def test_classifica_requisitorio(pdf_modelo_bytes):
    texto = _extrair_texto_pdf(pdf_modelo_bytes)
    assert eh_documento_requisitorio(texto) is True
    assert eh_documento_vinculo(texto) is False


def test_classifica_vinculo(pdf_vinculo_bytes):
    texto = _extrair_texto_pdf(pdf_vinculo_bytes)
    assert eh_documento_vinculo(texto) is True
    assert eh_documento_requisitorio(texto) is False


from baixar_requisitorios import casar_requisitorios_com_vinculos


def _req(ofreq, nome, doc, tipo):
    return {
        "ofreq": ofreq,
        "beneficiario": {"nome": nome, "doc": doc, "tipo_doc": tipo},
        "advogado": {"nome": "ADV " + nome, "cpf": "33394784068", "oab": "RJ1"},
    }


def test_casar_um_requisitorio_com_vinculo():
    reqs = [_req("2025.06478", "CARRARO", "28123344000107", "CNPJ")]
    vinc = {"2025.06478": "2025.06209-0"}
    dados = casar_requisitorios_com_vinculos(reqs, vinc, ["2025.06209-0"])
    assert dados["2025.06209-0"]["beneficiario_nome"] == "CARRARO"
    assert dados["2025.06209-0"]["beneficiario_doc"] == "28.123.344/0001-07"
    assert dados["2025.06209-0"]["advogado_cpf"] == "333.947.840-68"
    assert dados["2025.06209-0"]["status"] == "OK"


def test_casar_multiplos_precatorios_no_processo():
    reqs = [
        _req("2025.001", "AUTOR", "11111111111", "CPF"),
        _req("2025.002", "HONORARIOS LTDA", "22222222000122", "CNPJ"),
    ]
    vinc = {"2025.001": "2025.10-0", "2025.002": "2025.20-1"}
    dados = casar_requisitorios_com_vinculos(reqs, vinc, ["2025.10-0", "2025.20-1"])
    assert dados["2025.10-0"]["beneficiario_nome"] == "AUTOR"
    assert dados["2025.20-1"]["beneficiario_nome"] == "HONORARIOS LTDA"


def test_casar_precatorio_sem_vinculo_vira_revisar():
    reqs = []
    dados = casar_requisitorios_com_vinculos(reqs, {}, ["2025.99-9"])
    assert dados["2025.99-9"] == {"status": "REVISAR"}


def test_casar_requisitorio_sem_advogado():
    reqs = [{"ofreq": "2025.003", "beneficiario": {"nome": "X", "doc": "11111111111", "tipo_doc": "CPF"}, "advogado": None}]
    vinc = {"2025.003": "2025.30-0"}
    dados = casar_requisitorios_com_vinculos(reqs, vinc, ["2025.30-0"])
    assert dados["2025.30-0"]["advogado_nome"] is None
    assert dados["2025.30-0"]["status"] == "OK"


def test_atualizar_planilha_grava_status_revisar(xlsx_para_atualizar, tmp_path):
    saida = tmp_path / "saida.xlsx"
    atualizar_planilha(xlsx_para_atualizar, saida, {
        "2025.00001-0": {"status": "REVISAR"},
    })
    from openpyxl import load_workbook
    wb = load_workbook(saida)
    ws = wb.active
    assert ws.cell(row=1, column=19).value == "Status"      # cabecalho S
    assert ws.cell(row=2, column=19).value == "REVISAR"     # S
    assert ws.cell(row=2, column=14).value is None          # N vazio


def test_atualizar_planilha_grava_status_ok(xlsx_para_atualizar, tmp_path):
    saida = tmp_path / "saida.xlsx"
    atualizar_planilha(xlsx_para_atualizar, saida, {
        "2025.00001-0": {
            "beneficiario_nome": "TECHNE", "beneficiario_doc": "X",
            "advogado_nome": "A", "advogado_cpf": "B", "advogado_oab": "C",
            "status": "OK",
        },
    })
    from openpyxl import load_workbook
    wb = load_workbook(saida)
    ws = wb.active
    assert ws.cell(row=2, column=14).value == "TECHNE"
    assert ws.cell(row=2, column=19).value == "OK"


# ===== Timeout por processo (deadline do lote) =====
# Um único processo travado (await do Playwright sem timeout numa página wedged,
# ex: variante de PJe) NÃO pode congelar o lote inteiro.
import baixar_requisitorios as _br_timeout


def test_processar_com_timeout_estoura_marca_erro_navegacao(monkeypatch):
    """Processo que excede o deadline vira erro_navegacao e o wrapper RETORNA
    (não levanta) — assim o loop principal continua no próximo processo."""
    async def fake_travado(*a, **k):
        await asyncio.sleep(5)
        return {"status": "ok"}
    monkeypatch.setattr(_br_timeout, "processar_processo", fake_travado)
    resultado = asyncio.run(_br_timeout._processar_com_timeout(
        None, ["X"], "0001234-00.2020.8.19.0001", None, None, timeout_seg=0.1))
    assert resultado["status"] == "erro_navegacao"
    assert "timeout" in resultado["motivo"].lower()


def test_processar_com_timeout_passa_resultado_quando_rapido(monkeypatch):
    """Processo que termina dentro do deadline devolve o resultado intacto."""
    async def fake_rapido(*a, **k):
        return {"status": "ok", "arquivos": ["a.pdf"], "dados": {}}
    monkeypatch.setattr(_br_timeout, "processar_processo", fake_rapido)
    resultado = asyncio.run(_br_timeout._processar_com_timeout(
        None, ["X"], "0001234-00.2020.8.19.0001", None, None, timeout_seg=5))
    assert resultado["status"] == "ok"
    assert resultado["arquivos"] == ["a.pdf"]


def test_processar_com_timeout_propaga_login_expirado(monkeypatch):
    """Sessão expirada deve propagar pra main parar e salvar, não virar timeout."""
    async def fake_login_caiu(*a, **k):
        raise LoginExpiradoError("sessao caiu")
    monkeypatch.setattr(_br_timeout, "processar_processo", fake_login_caiu)
    with pytest.raises(LoginExpiradoError):
        asyncio.run(_br_timeout._processar_com_timeout(
            None, ["X"], "0001234-00.2020.8.19.0001", None, None, timeout_seg=5))


# ===== Fallback por conteúdo: requisitórios arquivados como "Petição" genérica =====
# Processo real 0110128-60.2015: requisitórios e vínculos DEPJU arquivados como
# "NNN - <docnum> - Petição" (rótulo genérico). A seleção por rótulo não casa nada,
# mas os CLASSIFICADORES reconhecem o conteúdo — então o fallback baixa nós genéricos
# e deixa o conteúdo decidir. Texto abaixo transcrito dos documentos reais.
from baixar_requisitorios import (
    eh_documento_requisitorio, eh_documento_vinculo,
    REGEX_OFREQ, REGEX_VINCULO, _indices_fallback, _faltam_vinculos,
    _todos_resolvidos,
)


def test_faltam_vinculos_todos_ligados_false():
    reqs = [{"ofreq": "2024.16124"}, {"ofreq": "2024.16126"}]
    vincs = {"2024.16124": "2025.09513-3", "2024.16126": "2025.09514-1"}
    assert _faltam_vinculos(reqs, vincs) is False


def test_faltam_vinculos_um_sem_vinculo_true():
    reqs = [{"ofreq": "2024.16124"}, {"ofreq": "2024.16126"}]
    vincs = {"2024.16124": "2025.09513-3"}  # falta o 16126
    assert _faltam_vinculos(reqs, vincs) is True


def test_faltam_vinculos_ofreq_none_ignorado():
    # requisitório sem OFREQ extraído não pode ser ligado — não dispara o fallback à toa
    assert _faltam_vinculos([{"ofreq": None}], {}) is False


def test_faltam_vinculos_lista_vazia_false():
    assert _faltam_vinculos([], {}) is False


# ===== _todos_resolvidos =====
def test_todos_resolvidos_lista_vazia_false():
    assert _todos_resolvidos([], [{"ofreq": "2024.1"}], {"2024.1": "2025.10-0"}) is False

def test_todos_resolvidos_completo_true():
    reqs = [{"ofreq": "2024.1"}, {"ofreq": "2024.2"}]
    vinc = {"2024.1": "2025.10-0", "2024.2": "2025.20-1"}
    assert _todos_resolvidos(["2025.10-0", "2025.20-1"], reqs, vinc) is True

def test_todos_resolvidos_parcial_false():
    reqs = [{"ofreq": "2024.1"}]
    vinc = {"2024.1": "2025.10-0", "2024.2": "2025.20-1"}
    assert _todos_resolvidos(["2025.10-0", "2025.20-1"], reqs, vinc) is False

def test_todos_resolvidos_modo_teste_false():
    reqs = [{"ofreq": "2024.1"}]
    vinc = {"2024.1": "2025.10-0"}
    assert _todos_resolvidos(["TESTE"], reqs, vinc) is False

def test_todos_resolvidos_vinculo_sem_requisitorio_false():
    assert _todos_resolvidos(["2025.10-0"], [], {"2024.1": "2025.10-0"}) is False

def test_todos_resolvidos_requisitorio_sem_vinculo_false():
    assert _todos_resolvidos(["2025.10-0"], [{"ofreq": "2024.1"}], {}) is False


# ===== precisa_processar: resume não re-lê o que já está no checkpoint =====
# Falsos erros de queda de sessão (erro_navegacao) não devem ser re-lidos a cada
# resume; ficam para um passe final de --apenas-erros.
from baixar_requisitorios import precisa_processar


def test_precisa_processar_novo_processo_true():
    assert precisa_processar("X", {}, apenas_erros=False) is True


def test_precisa_processar_normal_pula_ok():
    assert precisa_processar("X", {"X": {"status": "ok"}}, apenas_erros=False) is False


def test_precisa_processar_normal_pula_erro_navegacao():
    # MUDANÇA: no run normal, erro_navegacao NÃO é mais re-lido (evita re-leitura no resume)
    assert precisa_processar("X", {"X": {"status": "erro_navegacao"}}, apenas_erros=False) is False


def test_precisa_processar_normal_pula_sem_requisitorio():
    assert precisa_processar("X", {"X": {"status": "sem_requisitorio"}}, apenas_erros=False) is False


def test_precisa_processar_apenas_erros_refaz_nao_ok():
    assert precisa_processar("X", {"X": {"status": "erro_navegacao"}}, apenas_erros=True) is True


def test_precisa_processar_apenas_erros_pula_ok():
    assert precisa_processar("X", {"X": {"status": "ok"}}, apenas_erros=True) is False

_TEXTO_REQUISITORIO_0110128 = (
    "Estado do Rio de Janeiro Poder Judiciario Tribunal de Justica\n"
    "Cartorio da 15a Vara de Fazenda Publica\n"
    "Definitivo OFICIO No: 2024.16124/OFREQ\n"
    "OFICIO REQUISITORIO DE PAGAMENTO DE VERBA ALIMENTICIA\n"
    "I - TIPO DE REQUISICAO: Originaria\n"
    "Natureza: ALIMENTICIA\n"
    "II - ENTIDADE EXECUTADA Nome: ESTADO DO RIO DE JANEIRO\n"
    "CNPJ: 42.498.600/0001-71\n"
)

_TEXTO_VINCULO_0110128 = (
    "Departamento de Precatorios Judiciais - DEPJU\n"
    "Oficio DEPRE/DEPJU/HOLOS no AUT.2025.007878\n"
    "Processo Originario n. 0110128-60.2015.8.19.0001\n"
    "Oficio 2024.16126/OFREQ foi analisado pelo processo de analise "
    "00008054/2025 e gerou o precatorio 2025.09514-1\n"
)


def test_classificador_reconhece_requisitorio_formato_peticao():
    assert eh_documento_requisitorio(_TEXTO_REQUISITORIO_0110128) is True


def test_classificador_reconhece_vinculo_depju_formato_peticao():
    assert eh_documento_vinculo(_TEXTO_VINCULO_0110128) is True


def test_regex_ofreq_extrai_numero_do_requisitorio_real():
    m = REGEX_OFREQ.search(_TEXTO_REQUISITORIO_0110128)
    assert m is not None and m.group(1) == "2024.16124"


def test_regex_vinculo_extrai_ofreq_e_precatorio_real():
    m = REGEX_VINCULO.search(_TEXTO_VINCULO_0110128)
    assert m is not None
    assert m.group(1) == "2024.16126"
    assert m.group(2) == "2025.09514-1"


def test_indices_fallback_exclui_ja_tentados():
    assert _indices_fallback([1, 2, 3, 4], {2}, 10) == [1, 3, 4]


def test_indices_fallback_respeita_limite():
    assert _indices_fallback([1, 2, 3, 4, 5], set(), 2) == [1, 2]


def test_indices_fallback_preserva_ordem():
    assert _indices_fallback([5, 3, 1], set(), 10) == [5, 3, 1]


# ===== Early-stop do fallback =====
# Depois de achar o bloco de requisitórios/vínculos (varrendo do fim), parar após N
# downloads irrelevantes seguidos — evita baixar a cauda de petições antigas inúteis.
class _FakeNth:
    def nth(self, i):
        return ("el", i)


class _FakeVisu:
    def locator(self, sel):
        return _FakeNth()


def test_baixar_indices_early_stop_apos_misses(monkeypatch):
    """Para 3 misses após o último achado; não processa o resto."""
    relev = {10: False, 11: False, 12: True, 13: False, 14: False, 15: False, 16: True}
    processados = []

    async def fake_um(visualizador, el, idx, *a, **k):
        processados.append(idx)
        return relev[idx]
    monkeypatch.setattr(_br_timeout, "_baixar_e_classificar_um", fake_um)
    asyncio.run(_br_timeout._baixar_e_classificar_indices(
        _FakeVisu(), [10, 11, 12, 13, 14, 15, 16], None, [], {}, set(),
        log=lambda m: None, numero_processo=None, parar_apos_misses=3))
    # 10(F,pré) 11(F,pré) 12(T,reset) 13(m1) 14(m2) 15(m3->stop). 16 não é processado.
    assert processados == [10, 11, 12, 13, 14, 15]


def test_baixar_indices_misses_antes_de_achar_nao_param(monkeypatch):
    """Misses antes de qualquer achado NÃO disparam o early-stop (processa todos)."""
    processados = []

    async def fake_um(visualizador, el, idx, *a, **k):
        processados.append(idx)
        return False
    monkeypatch.setattr(_br_timeout, "_baixar_e_classificar_um", fake_um)
    asyncio.run(_br_timeout._baixar_e_classificar_indices(
        _FakeVisu(), [1, 2, 3, 4, 5], None, [], {}, set(),
        log=lambda m: None, numero_processo=None, parar_apos_misses=2))
    assert processados == [1, 2, 3, 4, 5]


def test_baixar_indices_sem_parar_apos_misses_processa_tudo(monkeypatch):
    """Sem parar_apos_misses (None, default do passo 1) processa tudo mesmo após misses."""
    relev = {1: True, 2: False, 3: False, 4: False, 5: False}
    processados = []

    async def fake_um(visualizador, el, idx, *a, **k):
        processados.append(idx)
        return relev[idx]
    monkeypatch.setattr(_br_timeout, "_baixar_e_classificar_um", fake_um)
    asyncio.run(_br_timeout._baixar_e_classificar_indices(
        _FakeVisu(), [1, 2, 3, 4, 5], None, [], {}, set(),
        log=lambda m: None, numero_processo=None))
    assert processados == [1, 2, 3, 4, 5]


def test_baixar_indices_goal_stop_para_ao_resolver(monkeypatch):
    """Para assim que todos os precatórios-alvo ficam resolvidos (não lê o resto)."""
    processados = []

    async def fake_um(visualizador, el, idx, pasta_temp, requisitorios, vinculos, *a, **k):
        processados.append(idx)
        if idx == 10:
            vinculos["2024.1"] = "2025.10-0"
            requisitorios.append({"ofreq": "2024.1"})
            return True
        return False
    monkeypatch.setattr(_br_timeout, "_baixar_e_classificar_um", fake_um)
    asyncio.run(_br_timeout._baixar_e_classificar_indices(
        _FakeVisu(), [10, 11, 12], None, [], {}, set(),
        log=lambda m: None, numero_processo=None,
        precatorios_alvo=["2025.10-0"]))
    assert processados == [10]  # parou logo após resolver o único alvo


def test_baixar_indices_goal_stop_segue_se_falta_alvo(monkeypatch):
    """Com 2 alvos e só 1 resolvido, NÃO para — segue lendo."""
    processados = []

    async def fake_um(visualizador, el, idx, pasta_temp, requisitorios, vinculos, *a, **k):
        processados.append(idx)
        if idx == 10:
            vinculos["2024.1"] = "2025.10-0"
            requisitorios.append({"ofreq": "2024.1"})
            return True
        return False
    monkeypatch.setattr(_br_timeout, "_baixar_e_classificar_um", fake_um)
    asyncio.run(_br_timeout._baixar_e_classificar_indices(
        _FakeVisu(), [10, 11, 12], None, [], {}, set(),
        log=lambda m: None, numero_processo=None,
        precatorios_alvo=["2025.10-0", "2025.20-1"]))
    assert processados == [10, 11, 12]  # nunca resolveu o 2º alvo -> leu tudo
