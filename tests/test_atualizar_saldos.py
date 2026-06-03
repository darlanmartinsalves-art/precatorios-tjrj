from atualizar_saldos import extrair_precatorios
from unittest.mock import MagicMock, patch
from atualizar_saldos import consultar_saldo

def _mock_response(status_code, json_data=None):
    m = MagicMock()
    m.status_code = status_code
    m.json.return_value = json_data or {}
    return m

def test_extrair_precatorios_retorna_lista_de_tuplas(xlsx_pequena):
    resultado = extrair_precatorios(xlsx_pequena)
    assert len(resultado) == 5
    assert resultado[0] == (2, "2025.09451-0")
    assert resultado[4] == (6, "2025.09470-6")

def test_extrair_precatorios_ignora_linhas_sem_numero_valido(xlsx_pequena, tmp_path):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_pequena)
    ws = wb.active
    ws.append(["", "INVALIDO", "", "", 0, "", 0, None])
    ws.append(["TJ", "2025.99999-9", "Comum", "", 2027, "", 0, None])
    saida = tmp_path / "modificada.xlsx"
    wb.save(saida)

    resultado = extrair_precatorios(saida)
    numeros = [n for _, n in resultado]
    assert "INVALIDO" not in numeros
    assert "2025.99999-9" in numeros

def test_consultar_saldo_retorna_float_em_caso_de_sucesso():
    session = MagicMock()
    session.get.return_value = _mock_response(200, {"Saldo": 13613821.78})
    resultado = consultar_saldo("2025.17049-6", session)
    assert resultado == 13613821.78

def test_consultar_saldo_404_retorna_nao_encontrado():
    session = MagicMock()
    session.get.return_value = _mock_response(404)
    resultado = consultar_saldo("9999.99999-9", session)
    assert resultado == "NAO_ENCONTRADO"

def test_consultar_saldo_null_retorna_sem_saldo():
    session = MagicMock()
    session.get.return_value = _mock_response(200, {"Saldo": None})
    resultado = consultar_saldo("2025.00001-1", session)
    assert resultado == "SEM_SALDO"

def test_consultar_saldo_retry_em_5xx():
    session = MagicMock()
    session.get.side_effect = [
        _mock_response(500),
        _mock_response(500),
        _mock_response(200, {"Saldo": 100.0}),
    ]
    with patch("atualizar_saldos.time.sleep"):
        resultado = consultar_saldo("2025.00001-1", session)
    assert resultado == 100.0
    assert session.get.call_count == 3

def test_consultar_saldo_erro_rede_apos_max_tentativas():
    session = MagicMock()
    session.get.side_effect = [_mock_response(500)] * 3
    with patch("atualizar_saldos.time.sleep"):
        resultado = consultar_saldo("2025.00001-1", session)
    assert resultado == "ERRO_REDE"


from atualizar_saldos import carregar_checkpoint, salvar_checkpoint

def test_checkpoint_inexistente_retorna_dict_vazio(tmp_path):
    caminho = tmp_path / "checkpoint.json"
    assert carregar_checkpoint(caminho) == {}

def test_salvar_e_carregar_checkpoint_preserva_dados(tmp_path):
    caminho = tmp_path / "checkpoint.json"
    dados = {"2025.09451-0": 59308.24, "2025.09452-8": "ERRO_REDE"}
    salvar_checkpoint(caminho, dados)
    assert carregar_checkpoint(caminho) == dados

def test_salvar_checkpoint_e_atomico(tmp_path):
    caminho = tmp_path / "checkpoint.json"
    salvar_checkpoint(caminho, {"a": 1})
    # após salvar, não deve sobrar arquivo .tmp
    assert not (tmp_path / "checkpoint.json.tmp").exists()
    assert caminho.exists()


from atualizar_saldos import escrever_xlsx
from openpyxl import load_workbook
import hashlib

def _md5(caminho):
    return hashlib.md5(caminho.read_bytes()).hexdigest()

def test_escrever_xlsx_preenche_coluna_H(xlsx_pequena, tmp_path):
    saida = tmp_path / "saida.xlsx"
    precatorios = [(2, "2025.09451-0"), (3, "2025.09452-8"), (4, "2025.09453-6")]
    resultados = {
        "2025.09451-0": 59308.24,
        "2025.09452-8": 83433.75,
        "2025.09453-6": "ERRO_REDE",  # não deve gravar
    }
    escrever_xlsx(xlsx_pequena, saida, precatorios, resultados)

    wb = load_workbook(saida)
    ws = wb.active
    assert ws.cell(row=2, column=8).value == 59308.24
    assert ws.cell(row=3, column=8).value == 83433.75
    assert ws.cell(row=4, column=8).value is None  # erro não gravado

def test_escrever_xlsx_aplica_formato_moeda(xlsx_pequena, tmp_path):
    saida = tmp_path / "saida.xlsx"
    escrever_xlsx(xlsx_pequena, saida, [(2, "2025.09451-0")], {"2025.09451-0": 100.0})
    wb = load_workbook(saida)
    ws = wb.active
    assert "#,##0.00" in ws.cell(row=2, column=8).number_format

def test_escrever_xlsx_nao_altera_entrada(xlsx_pequena, tmp_path):
    saida = tmp_path / "saida.xlsx"
    hash_antes = _md5(xlsx_pequena)
    escrever_xlsx(xlsx_pequena, saida, [(2, "2025.09451-0")], {"2025.09451-0": 100.0})
    hash_depois = _md5(xlsx_pequena)
    assert hash_antes == hash_depois
