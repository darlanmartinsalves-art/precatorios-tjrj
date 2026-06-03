import pytest
from openpyxl import Workbook
from pathlib import Path

@pytest.fixture
def xlsx_pequena(tmp_path):
    """Cria um xlsx temporário com 5 precatórios para testes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Precatorios"
    ws.append(["Tribunal", "Precatório", "Natureza", "Devedor", "Orçamento",
               "Processo Judicial", "Valor Histórico", "Saldo atualizado"])
    dados = [
        ("TJ", "2025.09451-0", "Alimentícia", "UERJ", 2027, "0230764-50.2018.8.19.0001", 50000, None),
        ("TJ", "2025.09452-8", "Alimentícia", "ESTADO", 2027, "0006013-75.2021.8.19.0001", 70000, None),
        ("TJ", "2025.09453-6", "Alimentícia", "IPERJ", 2027, "0027542-97.2014.8.19.0001", 300000, None),
        ("TJ", "2025.09462-5", "Comum", "ESTADO", 2027, "0000224-76.2021.8.19.0072", 100000, None),
        ("TJ", "2025.09470-6", "Comum", "ESTADO", 2027, "0123018-26.2018.8.19.0001", 350000, None),
    ]
    for d in dados:
        ws.append(d)
    caminho = tmp_path / "pequena.xlsx"
    wb.save(caminho)
    return caminho


@pytest.fixture
def pdf_modelo_bytes():
    """Retorna o conteúdo binário do PDF de referência (Garrastazu/TECHNE)."""
    caminho = Path(__file__).parent / "fixtures" / "garrastazu_modelo.pdf"
    return caminho.read_bytes()


@pytest.fixture
def pdf_modelo_path():
    """Retorna o Path do PDF de referência."""
    return Path(__file__).parent / "fixtures" / "garrastazu_modelo.pdf"


@pytest.fixture
def pdf_vinculo_bytes():
    """Conteúdo binário do ofício DEPRE/DEPJU que liga OFREQ -> precatório."""
    caminho = Path(__file__).parent / "fixtures" / "vinculo_depre_modelo.pdf"
    return caminho.read_bytes()


@pytest.fixture
def xlsx_com_saldos(tmp_path):
    """xlsx com 6 precatórios, saldos variados, alguns processos duplicados."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Tribunal", "Precatório", "Natureza", "Devedor", "Orçamento",
               "Processo Judicial", "Valor Histórico", "Saldo atualizado"])
    # (precatorio, processo, saldo)
    dados = [
        ("TJ", "2025.00001-0", "Comum", "ESTADO", 2027, "0000001-00.2020.8.19.0001", 100, 100000.00),
        ("TJ", "2025.00002-1", "Comum", "ESTADO", 2027, "0000002-00.2020.8.19.0001", 100, 250000.00),
        ("TJ", "2025.00003-2", "Comum", "ESTADO", 2027, "0000003-00.2020.8.19.0001", 100, 500000.00),
        ("TJ", "2025.00004-3", "Comum", "ESTADO", 2027, "0000003-00.2020.8.19.0001", 100, 300000.00),  # processo duplicado
        ("TJ", "2025.00005-4", "Comum", "ESTADO", 2027, "0000004-00.2020.8.19.0001", 100, 199999.99),
        ("TJ", "2025.00006-5", "Comum", "ESTADO", 2027, "0000005-00.2020.8.19.0001", 100, None),
    ]
    for d in dados:
        ws.append(d)
    caminho = tmp_path / "saldos.xlsx"
    wb.save(caminho)
    return caminho


@pytest.fixture
def xlsx_para_atualizar(tmp_path):
    """xlsx pequeno com 3 precatórios para testar atualizar_planilha."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Tribunal", "Precatório", "Natureza", "Devedor", "Orçamento",
               "Processo Judicial", "Valor Histórico", "Saldo atualizado",
               "Cedente", "CPF", "Celular", "E-mail", "OBSERVAÇÃO"])
    ws.append(("TJ", "2025.00001-0", "Comum", "ESTADO", 2027,
               "0000001-00.2020.8.19.0001", 100, 250000.00,
               "ALGUEM", "111.111.111-11", "(21) 11111-1111", "a@a.com", ""))
    ws.append(("TJ", "2025.00002-1", "Comum", "ESTADO", 2027,
               "0000002-00.2020.8.19.0001", 100, 500000.00,
               "OUTRO", "222.222.222-22", "(21) 22222-2222", "b@b.com", ""))
    ws.append(("TJ", "2025.00003-2", "Comum", "ESTADO", 2027,
               "0000003-00.2020.8.19.0001", 100, 100000.00,
               "MAIS UM", "333.333.333-33", "", "c@c.com", ""))
    caminho = tmp_path / "planilha_atualizavel.xlsx"
    wb.save(caminho)
    return caminho
