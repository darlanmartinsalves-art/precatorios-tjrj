"""Baixa ofícios requisitórios em PDF do portal TJRJ para precatórios filtrados."""
import re
import json
import asyncio
import sys
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook
from pypdf import PdfReader
from playwright.async_api import async_playwright, BrowserContext, Page
import seletores


class LoginExpiradoError(Exception):
    """Disparado quando a sessão expirou durante a navegação."""
    pass


class ProcessoPjeError(Exception):
    """Disparado quando o processo pertence ao sistema PJe (sem acesso)."""
    pass

REGEX_CNJ = re.compile(r"^(\d{7}-\d{2}\.\d{4})\.8\.19\.(\d{4})$")


def split_cnj(numero):
    """Separa um número CNJ em (prefixo, sufixo) para os 2 campos do form.

    Ex: "0156129-30.2020.8.19.0001" -> ("0156129-30.2020", "0001")

    Lança ValueError se o número não casa com o padrão CNJ TJRJ.
    """
    if not numero or not isinstance(numero, str):
        raise ValueError(f"Número CNJ inválido: {numero!r}")
    m = REGEX_CNJ.match(numero.strip())
    if not m:
        raise ValueError(f"Não casa com padrão CNJ TJRJ: {numero!r}")
    return m.group(1), m.group(2)


CHARS_INVALIDOS = r'\/:*?"<>|'


def sanitizar_nome(nome):
    """Sanitiza um nome para uso seguro como nome de arquivo no Windows.

    - Substitui caracteres inválidos (/\\:*?"<>|) por _
    - Colapsa múltiplos espaços em um único
    - Faz strip de espaços nas pontas
    - Limita a 100 caracteres
    """
    if not nome:
        return ""
    saida = "".join("_" if c in CHARS_INVALIDOS else c for c in nome)
    saida = re.sub(r"\s+", " ", saida).strip()
    return saida[:100]


REGEX_BENEFICIARIO_NOME = re.compile(
    r"III\s*[-–]\s*BENEFICI[ÁA]RIO.*?Nome:\s*(.+?)(?:\n|CNPJ|CPF)",
    re.DOTALL | re.IGNORECASE,
)
REGEX_BENEFICIARIO_DOC = re.compile(
    r"III\s*[-–]\s*BENEFICI[ÁA]RIO.*?(CPF|CNPJ):\s*([\d./\-]+)",
    re.DOTALL | re.IGNORECASE,
)


def _extrair_texto_pdf(pdf_bytes):
    """Helper: extrai todo o texto de um PDF bytes. Retorna string vazia se falhar."""
    if not pdf_bytes:
        return ""
    try:
        reader = PdfReader(BytesIO(pdf_bytes))
    except Exception:
        return ""
    texto = ""
    for pagina in reader.pages:
        try:
            texto += pagina.extract_text() + "\n"
        except Exception:
            continue
    return texto


def extrair_beneficiario_completo(pdf_bytes):
    """Extrai nome, doc e tipo_doc do beneficiário (seção III) do PDF.

    Retorna dict {"nome": str, "doc": str (só dígitos), "tipo_doc": "CPF"|"CNPJ"}
    ou None se não conseguir.
    """
    texto = _extrair_texto_pdf(pdf_bytes)
    if not texto:
        return None
    m_nome = REGEX_BENEFICIARIO_NOME.search(texto)
    if not m_nome:
        return None
    m_doc = REGEX_BENEFICIARIO_DOC.search(texto)
    if not m_doc:
        return None
    tipo = m_doc.group(1).upper()
    doc_raw = re.sub(r"\D", "", m_doc.group(2))
    return {"nome": m_nome.group(1).strip(), "doc": doc_raw, "tipo_doc": tipo}


def extrair_beneficiario(pdf_bytes):
    """Função legada: retorna apenas o nome do beneficiário. Use extrair_beneficiario_completo
    para obter também CPF/CNPJ.
    """
    completo = extrair_beneficiario_completo(pdf_bytes)
    return completo["nome"] if completo else None


REGEX_ADVOGADO = re.compile(
    r"VII\s*[-–]\s*ADVOGADO[S]?\s*DO\s*BENEFICI[ÁA]RIO.*?"
    r"([A-Z]{2}\d+)\s*[-–]\s*(.+?)\s*[-–]\s*(\d{11})",
    re.DOTALL | re.IGNORECASE,
)


def extrair_advogado(pdf_bytes):
    """Extrai o primeiro advogado (seção VII) do PDF.

    Formato esperado: "OAB - NOME - CPF" onde OAB é tipo "RJ185918".

    Retorna dict {"nome": str, "cpf": str (só dígitos), "oab": str}
    ou None se não conseguir.
    """
    texto = _extrair_texto_pdf(pdf_bytes)
    if not texto:
        return None
    m = REGEX_ADVOGADO.search(texto)
    if not m:
        return None
    return {
        "oab": m.group(1).strip(),
        "nome": m.group(2).strip(),
        "cpf": m.group(3).strip(),
    }


REGEX_OFREQ = re.compile(r"(\d{4}\.\d+)\s*/\s*OFREQ", re.IGNORECASE)


def extrair_numero_ofreq(pdf_bytes):
    """Extrai o número OFREQ (ex: '2025.14235') do texto do PDF, ou None."""
    texto = _extrair_texto_pdf(pdf_bytes)
    if not texto:
        return None
    m = REGEX_OFREQ.search(texto)
    return m.group(1) if m else None


REGEX_VINCULO = re.compile(
    r"Of[íi]cio\s+(\d{4}\.\d+)\s*/\s*OFREQ.*?"
    r"gerou\s+o\s+precat[óo]rio\s+(\d{4}\.\d+-\d+)",
    re.DOTALL | re.IGNORECASE,
)


def extrair_vinculo_ofreq_precatorio(pdf_bytes):
    """Do ofício DEPRE/DEPJU, retorna (ofreq, precatorio) ou None.

    Ex: 'Ofício 2025.06478/OFREQ ... gerou o precatório 2025.06209-0'
        -> ('2025.06478', '2025.06209-0')
    """
    texto = _extrair_texto_pdf(pdf_bytes)
    if not texto:
        return None
    m = REGEX_VINCULO.search(texto)
    return (m.group(1), m.group(2)) if m else None


REGEX_EH_REQUISITORIO = re.compile(r"OF[ÍI]CIO\s+REQUISIT[ÓO]RIO", re.IGNORECASE)
REGEX_EH_VINCULO = re.compile(r"gerou\s+o\s+precat[óo]rio", re.IGNORECASE)


def eh_documento_requisitorio(texto):
    """True se o texto é de um ofício requisitório (tem dados de beneficiário)."""
    return bool(texto) and bool(REGEX_EH_REQUISITORIO.search(texto))


def eh_documento_vinculo(texto):
    """True se o texto é o ofício DEPRE/DEPJU que liga OFREQ ao precatório."""
    return bool(texto) and bool(REGEX_EH_VINCULO.search(texto))


def gerar_nome_arquivo(precatorio, beneficiario, pasta):
    """Gera um Path único para o arquivo de saída.

    Formato base: "{precatorio} - {beneficiario sanitizado}.pdf"
    Se já existir, adiciona " (2)", " (3)" etc. até achar nome livre.
    """
    pasta = Path(pasta)
    base = f"{precatorio} - {sanitizar_nome(beneficiario)}"
    candidato = pasta / f"{base}.pdf"
    if not candidato.exists():
        return candidato
    ordinal = 2
    while True:
        candidato = pasta / f"{base} ({ordinal}).pdf"
        if not candidato.exists():
            return candidato
        ordinal += 1


REGEX_PRECATORIO = re.compile(r"^\d{4}\.\d+-\d+$")
REGEX_PROCESSO_CNJ = re.compile(r"^\d{7}-\d{2}\.\d{4}\.8\.19\.\d{4}$")


def filtrar_precatorios(caminho_xlsx, saldo_minimo):
    """Lê a planilha e retorna lista de (precatorio, processo) com saldo >= mínimo.

    Filtros aplicados:
    - Coluna B casa com padrão de precatório
    - Coluna F casa com padrão CNJ TJRJ
    - Coluna H é numérico e >= saldo_minimo
    """
    wb = load_workbook(caminho_xlsx, read_only=True, data_only=True)
    ws = wb.active
    resultado = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 8:
            continue
        precatorio = row[1]
        processo = row[5]
        saldo = row[7]
        if not isinstance(precatorio, str) or not REGEX_PRECATORIO.match(precatorio):
            continue
        if not isinstance(processo, str) or not REGEX_PROCESSO_CNJ.match(processo.strip()):
            continue
        if not isinstance(saldo, (int, float)):
            continue
        if saldo < saldo_minimo:
            continue
        resultado.append((precatorio, processo.strip()))
    wb.close()
    return resultado


def carregar_checkpoint_dl(caminho):
    """Carrega checkpoint JSON. Retorna {} se arquivo não existir."""
    caminho = Path(caminho)
    if not caminho.exists():
        return {}
    with open(caminho, "r", encoding="utf-8") as f:
        return json.load(f)


def salvar_checkpoint_dl(caminho, dados):
    """Salva checkpoint atomicamente: escreve .tmp e renomeia."""
    caminho = Path(caminho)
    tmp = caminho.with_suffix(caminho.suffix + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    tmp.replace(caminho)


PROJETO_DIR = Path(__file__).parent
USER_DATA_DIR = PROJETO_DIR / "user_data_chromium"
SCREENSHOTS_DIR = PROJETO_DIR / "screenshots_erro"
CHECKPOINT_PATH = PROJETO_DIR / "downloads_state.json"
ERROS_CSV = PROJETO_DIR / "erros_download.csv"

URL_BASE = "https://www3.tjrj.jus.br/portalservicos/"
URL_CONSULTA = "https://www3.tjrj.jus.br/portalservicos/#/consproc/consultaportal"


async def abrir_context(playwright, headless=False):
    """Abre BrowserContext persistente lançando o Edge do sistema via PIPE.

    Usa --remote-debugging-pipe (transporte padrão do launch), NÃO uma porta TCP de
    debug — assim evita a heurística do antivírus corporativo (Bitdefender) que
    bloqueia o tráfego CDP em porta de debug (técnica de infostealer).
    A sessão de login persiste em user_data_chromium/.

    Retorna o context. O caller deve fechar com `await context.close()`.
    """
    USER_DATA_DIR.mkdir(parents=True, exist_ok=True)
    context = await playwright.chromium.launch_persistent_context(
        user_data_dir=str(USER_DATA_DIR),
        channel="msedge",
        headless=headless,
        accept_downloads=True,
        viewport={"width": 1400, "height": 900},
        args=["--no-first-run", "--no-default-browser-check"],
    )
    return context


def _parece_logado_portal(url):
    """True se a URL indica sessão ativa no Portal de Serviços.

    Sem sessão, o portal (www3.../portalservicos/) redireciona para o site PÚBLICO
    www.tjrj.jus.br — que não tem campo de senha e enganava a checagem antiga.
    Logado de verdade = continua em 'portalservicos' e não é a tela de login.
    """
    u = (url or "").lower()
    return "portalservicos" in u and "login" not in u


async def verificar_sessao(context):
    """Verifica se há sessão ativa no Portal de Serviços abrindo a URL base.

    Retorna True só se, após o load, continuamos no portalservicos (sessão viva)
    e sem campo de senha. Se redirecionou pro site público, retorna False.
    """
    page = context.pages[0] if context.pages else await context.new_page()
    # NÃO usar networkidle: o portal é um SPA que mantém conexões abertas e o
    # networkidle frequentemente estoura 30s sem nunca ficar ocioso. O resto do
    # código usa domcontentloaded; aqui também, tolerante a timeout + breve settle.
    try:
        await page.goto(URL_BASE, wait_until="domcontentloaded", timeout=30000)
    except Exception:
        pass
    await asyncio.sleep(3)  # deixar o SPA assentar/redirecionar antes de checar
    if not _parece_logado_portal(page.url):
        return False
    senha_visivel = await page.locator('input[type="password"]').count()
    return senha_visivel == 0


async def login_interativo(context):
    """Abre o portal e aguarda o usuário logar manualmente.

    Bloqueia até o usuário pressionar ENTER no terminal.
    """
    page = context.pages[0] if context.pages else await context.new_page()
    await page.goto(URL_BASE, wait_until="domcontentloaded")
    print("\n" + "=" * 60)
    print(" LOGIN NECESSÁRIO")
    print(" Faça login no portal TJRJ no navegador que abriu.")
    print(" Quando estiver logado, pressione ENTER aqui.")
    print("=" * 60)
    input(" Pressione ENTER após logar: ")
    if await verificar_sessao(context):
        print(" Sessão confirmada. Continuando...")
        return True
    print(" Sessão ainda não confirmada. Saindo.", file=sys.stderr)
    return False


async def aguardar_login_no_browser(context, timeout_seg=900):
    """Abre o portal e ESPERA (faz polling) o usuário logar no Edge que abriu.

    Não usa terminal (input) — detecta o login sozinho observando a página. Assim
    funciona mesmo rodando em background. Considera logado quando a página está no
    Portal de Serviços (URL contém 'portalservicos') e não há campo de senha.
    Retorna True se logou dentro do timeout, False caso contrário.
    """
    page = context.pages[0] if context.pages else await context.new_page()
    try:
        await page.goto(URL_BASE, wait_until="domcontentloaded", timeout=30000)
    except Exception:
        pass
    print("\n" + "=" * 60)
    print(" LOGIN NECESSÁRIO")
    print(" Faça login no Edge que abriu: Processo Eletrônico ->")
    print(" Portal de Serviços -> CPF/senha/código -> perfil Advogado.")
    print(f" Aguardando você logar (até {timeout_seg // 60} min)...")
    print("=" * 60, flush=True)
    i = 0
    for _ in range(max(1, timeout_seg // 5)):
        await asyncio.sleep(5)
        i += 1
        try:
            # Checar TODAS as abas — o login pode abrir o portal em nova aba
            for p in list(context.pages):
                if _parece_logado_portal(p.url):
                    if await p.locator('input[type="password"]').count() == 0:
                        print(" Login detectado! Continuando...", flush=True)
                        await asyncio.sleep(2)
                        return True
            # Diagnóstico (a cada ~30s): mostra as abas vistas e por que não casaram,
            # pra distinguir "não logou" de "logou mas detecção falhou".
            if i % 6 == 0:
                infos = []
                for p in list(context.pages):
                    try:
                        tem_senha = await p.locator('input[type="password"]').count()
                    except Exception:
                        tem_senha = "?"
                    infos.append(f"{p.url} (portal={_parece_logado_portal(p.url)}, senha={tem_senha})")
                print(f" [aguardando login] abas vistas:\n   " + "\n   ".join(infos), flush=True)
        except Exception:
            continue
    print(" Tempo de login esgotado.", file=sys.stderr)
    return False


async def _test_login_manual():
    """Função de teste manual: abre o browser, valida sessão (ou pede login)."""
    async with async_playwright() as pw:
        context = await abrir_context(pw, headless=False)
        try:
            if await verificar_sessao(context):
                print("✅ Já estava logado!")
            else:
                logou = await login_interativo(context)
                if logou:
                    print("✅ Login feito com sucesso.")
                else:
                    print("❌ Falha no login.")
            input("Pressione ENTER para fechar...")
        finally:
            await context.close()


def formatar_doc(doc, tipo):
    """Formata um CPF ou CNPJ.

    CPF: '33394784068' -> '333.947.840-68'
    CNPJ: '50737766000121' -> '50.737.766/0001-21'

    Se já estiver formatado ou tamanho não bater, retorna o doc original.
    """
    if not doc:
        return doc
    digitos = re.sub(r"\D", "", doc)
    if tipo == "CPF" and len(digitos) == 11:
        return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"
    if tipo == "CNPJ" and len(digitos) == 14:
        return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"
    return doc


COLUNAS_NOVAS = {
    14: "Beneficiário Nome",   # N
    15: "Beneficiário Doc",    # O
    16: "Advogado Nome",       # P
    17: "Advogado CPF",        # Q
    18: "Advogado OAB",        # R
    19: "Status",              # S
}


def atualizar_planilha(caminho_entrada, caminho_saida, dados):
    """Carrega planilha de entrada, atualiza colunas N-R, salva em saída.

    `dados` é dict {precatorio: {"beneficiario_nome", "beneficiario_doc",
                                  "advogado_nome", "advogado_cpf", "advogado_oab"}}

    Entrada permanece inalterada.
    """
    wb = load_workbook(caminho_entrada)
    ws = wb.active

    # Cabeçalhos novos (linha 1)
    for col, header in COLUNAS_NOVAS.items():
        ws.cell(row=1, column=col).value = header

    # Mapear precatório -> linha
    mapa = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row and len(row) > 1 and isinstance(row[1], str):
            mapa[row[1]] = row_idx

    # Preencher
    for precatorio, info in dados.items():
        row_idx = mapa.get(precatorio)
        if row_idx is None:
            continue
        ws.cell(row=row_idx, column=14).value = info.get("beneficiario_nome")
        ws.cell(row=row_idx, column=15).value = info.get("beneficiario_doc")
        ws.cell(row=row_idx, column=16).value = info.get("advogado_nome")
        ws.cell(row=row_idx, column=17).value = info.get("advogado_cpf")
        ws.cell(row=row_idx, column=18).value = info.get("advogado_oab")
        ws.cell(row=row_idx, column=19).value = info.get("status")

    wb.save(caminho_saida)
    wb.close()




CDP_PORTA_PADRAO = 9223  # Edge (Chrome usaria 9222 mas antivirus corporativo pode bloquear)


async def conectar_chrome_cdp(porta=CDP_PORTA_PADRAO):
    """Conecta no Edge/Chrome aberto via CDP. Retorna (playwright, browser, context).

    Pré-requisito: usuário deve ter aberto Edge com:
      msedge.exe --remote-debugging-port=9223 --user-data-dir=C:\\temp\\edge_debug

    Levanta RuntimeError se não conseguir conectar.
    """
    pw = await async_playwright().start()
    try:
        # Usar 127.0.0.1 (não 'localhost'): localhost resolve para ::1 (IPv6) primeiro,
        # e o Edge debug escuta só em IPv4 -> ECONNREFUSED ::1.
        browser = await pw.chromium.connect_over_cdp(f"http://127.0.0.1:{porta}")
    except Exception as e:
        await pw.stop()
        raise RuntimeError(
            f"Não foi possível conectar ao browser em 127.0.0.1:{porta}.\n"
            f"Erro: {e}\n"
            f"O Edge debug provavelmente NÃO está aberto. Rode:\n"
            f"  .\\iniciar_edge_debug.ps1\n"
            f"e faça login no Portal de Serviços (perfil Advogado)."
        )
    if not browser.contexts:
        await browser.close()
        await pw.stop()
        raise RuntimeError("Browser não tem nenhum context. Reabra o Edge.")
    context = browser.contexts[0]
    return pw, browser, context


async def _test_cdp_manual():
    """Função de teste manual: conecta no Chrome, lista URLs abertas."""
    pw = None
    browser = None
    try:
        pw, browser, context = await conectar_chrome_cdp()
        print(f"Conectado! Pages abertas:")
        for i, page in enumerate(context.pages):
            print(f"  [{i}] {page.url}")
        print(f"Total: {len(context.pages)} page(s)")
        await asyncio.sleep(1)
    finally:
        if browser:
            await browser.close()
        if pw:
            await pw.stop()


async def _erro_iframe_ausente(page, numero_processo):
    """Decide qual exceção lançar quando o iframe de consulta não carrega.

    Sessão expirada faz o portal redirecionar pra tela de login — o iframe
    'consultaprocessual' some. Nesse caso retorna LoginExpiradoError (sinal pro
    main() parar e salvar). Senão, RuntimeError (erro de navegação daquele processo).
    """
    if await page.locator(seletores.INDICADOR_LOGIN_EXPIRADO).count() > 0:
        return LoginExpiradoError(f"Sessão expirada ao consultar {numero_processo}")
    return RuntimeError(
        f"Iframe interno '{seletores.INNER_FRAME_URL_FRAGMENT}' não encontrado"
    )


async def _achar_frame_consulta(page, timeout_ms=15000):
    """Retorna o iframe interno da página de Consulta Processual (que tem os elementos reais).

    Faz poll por até timeout_ms aguardando o iframe aparecer.
    """
    intervalo = 0.5
    decorrido = 0.0
    while decorrido * 1000 < timeout_ms:
        for fr in page.frames:
            if seletores.INNER_FRAME_URL_FRAGMENT in fr.url:
                return fr
        await asyncio.sleep(intervalo)
        decorrido += intervalo
    return None


async def _solicitar_acesso_se_aparecer(page):
    """Se o modal de Solicitação de Acesso aparecer (advogado não vinculado),
    preenche o campo Motivo com qualquer texto e clica 'Visualizar Processo'.

    Usa o caminho SOLICITAÇÃO DE ACESSO (mais formal, conforme Resolução 121 CNJ).

    Retorna True se conseguiu interagir com o modal, False se não estava presente.
    """
    for fr in page.frames:
        try:
            campo = fr.locator(seletores.MODAL_ACESSO_MOTIVO)
            if await campo.count() == 0:
                continue
            # Preencher via JS (não precisa estar visível)
            try:
                await campo.first.evaluate(f"(el) => {{ el.value = {seletores.MOTIVO_PADRAO!r}; el.dispatchEvent(new Event('input', {{bubbles:true}})); el.dispatchEvent(new Event('change', {{bubbles:true}})); }}")
            except Exception:
                try:
                    await campo.first.fill(seletores.MOTIVO_PADRAO)
                except Exception:
                    continue
            await asyncio.sleep(0.5)
            # Clicar via JS (ignora overlays)
            btn = fr.locator(seletores.MODAL_ACESSO_BOTAO_OK)
            if await btn.count() == 0:
                continue
            try:
                await btn.first.evaluate("(el) => el.click()")
                await asyncio.sleep(2)
                return True
            except Exception:
                continue
        except Exception:
            continue
    return False


async def _detectar_processo_pje(page):
    """Detecta se o modal 'Mensagem Processo do PJe' está visível.
    Se sim, fecha o modal e retorna True (processo deve ser pulado).
    """
    for fr in page.frames:
        try:
            modal = fr.locator(seletores.MODAL_PJE)
            if await modal.count() > 0 and await modal.first.is_visible():
                # Fechar o modal
                fechar = fr.locator(seletores.MODAL_PJE_BOTAO_FECHAR).first
                if await fechar.count() > 0:
                    try:
                        await fechar.click(force=True, timeout=5000)
                    except Exception:
                        await fechar.evaluate("(el) => el.click()")
                    await asyncio.sleep(1)
                return True
        except Exception:
            continue
    return False


async def _resolver_alterar_perfil(context):
    """Se aparecer modal 'Alterar Perfil', seleciona 'Advogado' e clica Entrar."""
    for p in context.pages:
        if "tjrj.jus.br" not in p.url:
            continue
        for fr in p.frames:
            try:
                modal = fr.locator(seletores.MODAL_ALTERAR_PERFIL)
                if await modal.count() == 0:
                    continue
                if not await modal.first.is_visible():
                    continue
                # Clicar no dropdown para abrir as opções
                dropdown = fr.locator(seletores.MODAL_ALTERAR_PERFIL_DROPDOWN).first
                if await dropdown.is_visible():
                    await dropdown.click()
                    await asyncio.sleep(1)
                # Clicar em "Advogado"
                opcao = fr.locator(seletores.MODAL_ALTERAR_PERFIL_OPCAO_ADVOGADO).first
                if await opcao.is_visible():
                    await opcao.click()
                    await asyncio.sleep(0.5)
                # Clicar em Entrar
                btn = fr.locator(seletores.MODAL_ALTERAR_PERFIL_BOTAO_ENTRAR).first
                if await btn.is_visible():
                    await btn.click()
                    await asyncio.sleep(2)
                return True
            except Exception:
                continue
    return False


async def _clicar_prolongar_sessao(context):
    """Procura modal 'Aviso de sessão inativa' em todas as abas TJRJ e clica 'Prolongar sessão'.

    Também REMOVE qualquer modal-backdrop órfão que esteja interceptando cliques.

    Retorna o número de abas onde tomou alguma ação.
    """
    cliques = 0
    for p in context.pages:
        if "tjrj.jus.br" not in p.url:
            continue
        for fr in p.frames:
            try:
                # 1. Procurar modal de sessão visível e clicar Prolongar via JS
                modal = fr.locator(seletores.MODAL_SESSAO_INATIVA)
                if await modal.count() > 0 and await modal.first.is_visible():
                    btn = fr.locator(seletores.MODAL_SESSAO_PROLONGAR)
                    if await btn.count() > 0:
                        # Clicar via JS (ignora overlays)
                        try:
                            await btn.first.evaluate("(el) => el.click()")
                            cliques += 1
                            await asyncio.sleep(0.5)
                        except Exception:
                            pass

                # 2. Remover qualquer modal-backdrop órfão que possa estar interceptando
                # (acontece quando o modal de sessão fecha mas o backdrop fica preso)
                removidos = await fr.evaluate("""
                    () => {
                        const backdrops = document.querySelectorAll('.modal-backdrop');
                        let count = 0;
                        backdrops.forEach(b => {
                            // Verificar se o modal pai está oculto
                            const modais = document.querySelectorAll('.modal.show');
                            if (modais.length === 0) {
                                b.remove();
                                count++;
                            }
                        });
                        return count;
                    }
                """)
                if removidos:
                    cliques += 1
            except Exception:
                pass
    return cliques


async def consultar_processo(context, numero_processo, debug=False):
    """Navega de Consulta Processual até abrir o Visualizador do processo.

    Retorna a Page do visualizador (nova aba), ou None se o processo não foi
    encontrado ou se ocorreu erro recuperável (ex: sem acesso).

    Levanta LoginExpiradoError se detectar que a sessão expirou.
    """
    def log(msg):
        if debug:
            print(f"  [consultar] {msg}")

    # Antes de tudo, garantir que a sessão está viva e que perfil está selecionado
    await _clicar_prolongar_sessao(context)
    await _resolver_alterar_perfil(context)
    log("prolongar_sessao + alterar_perfil OK")

    # Reusar uma aba do portalservicos JÁ existente. Logo após o login, o SPA abre
    # dashboard e depois roteia para consultaportal — mas leva alguns segundos. Criar
    # uma aba NOVA nesse momento não carrega o iframe 'consultaprocessual' (causa do
    # erro_navegacao no 1º processo após o login). Então esperamos a aba do portal
    # surgir e a reusamos; só criamos aba nova como último recurso.
    page = None
    for _ in range(12):  # ~24s aguardando a aba do portal aparecer/rotear
        for p in context.pages:
            if "consultaportal" in p.url or "consultaprocessual" in p.url:
                page = p
                break
        if page is None:
            for p in context.pages:  # fallback: qualquer aba do portalservicos (ex: dashboard)
                if "portalservicos" in p.url:
                    page = p
                    break
        if page is not None:
            break
        await asyncio.sleep(2)
    if page is None:
        page = await context.new_page()

    log(f"page={page.url[:80]}")
    # Navegar pra consulta processual. NÃO usar page.reload() — redireciona pra alterar-perfil.
    # Cold-start pós-login: o iframe 'consultaprocessual' pode demorar a surgir no 1º
    # processo depois do login (o login abre várias abas e o SPA ainda assenta). Por isso
    # re-navegamos e procuramos com timeout maior, por até 3 tentativas, fechando o modal
    # de sessão inativa antes de cada busca. Só erra se o iframe não aparecer em nenhuma.
    inner = None
    for tentativa in range(3):
        await page.goto(seletores.URL_CONSULTA, wait_until="domcontentloaded", timeout=30000)
        await asyncio.sleep(2)
        log(f"apos goto (tentativa {tentativa + 1}), page={page.url[:80]}")
        await _clicar_prolongar_sessao(context)  # fecha modal de sessão / backdrop órfão
        inner = await _achar_frame_consulta(page, timeout_ms=30000)
        if inner is not None:
            break
        log(f"iframe nao apareceu (tentativa {tentativa + 1}) — re-tentando")
        await asyncio.sleep(3)
    if inner is None:
        raise await _erro_iframe_ausente(page, numero_processo)
    log(f"inner={inner.url[:100]}")

    # Se iframe está em estado diferente de consultaportal, forçar reset
    if "consultaportal" not in inner.url:
        inner_url = "https://www3.tjrj.jus.br/consultaprocessual/#/consultaportal"
        try:
            await inner.evaluate(f"window.location.replace('{inner_url}')")
        except Exception as e:
            log(f"erro no evaluate reset: {e}")
        await asyncio.sleep(3)
        inner = await _achar_frame_consulta(page, timeout_ms=15000)
        if inner is None:
            log("inner sumiu apos reset")
            return None
        log(f"inner apos reset={inner.url[:100]}")

    # Esperar form aparecer
    try:
        await inner.locator(seletores.RADIO_UNICA).wait_for(state="visible", timeout=15000)
        log("form RADIO_UNICA visivel")
    except Exception as e:
        log(f"form NAO apareceu: {e}")
        if await page.locator(seletores.INDICADOR_LOGIN_EXPIRADO).count() > 0:
            raise LoginExpiradoError(f"Sessão expirada ao consultar {numero_processo}")
        return None

    # Fechar modal de sessão inativa se aparecer (intercepta cliques)
    await _clicar_prolongar_sessao(context)

    # Preencher CNJ
    prefix, sufix = split_cnj(numero_processo)
    await inner.locator(seletores.RADIO_UNICA).check()
    await inner.locator(seletores.CAMPO_CNJ_PREFIX).fill(prefix)
    await inner.locator(seletores.CAMPO_CNJ_SUFIX).fill(sufix)

    # Fechar modal novamente se aparecer
    await _clicar_prolongar_sessao(context)

    # Clicar pesquisar (pegar botão visível — há vários por aba)
    botoes = await inner.locator(seletores.BOTAO_PESQUISAR).all()
    log(f"botoes pesquisar encontrados: {len(botoes)}")
    clicado = False
    for idx_b, b in enumerate(botoes):
        if await b.is_visible():
            log(f"  botao {idx_b} visivel, tentando click")
            for tentativa in range(3):
                try:
                    await b.click(timeout=8000)
                    clicado = True
                    log(f"  clicou (tentativa {tentativa+1})")
                    break
                except Exception as e:
                    log(f"  click falhou (tentativa {tentativa+1}): {str(e)[:100]}")
                    # Fechar qualquer modal
                    await _clicar_prolongar_sessao(context)
                    await asyncio.sleep(1)
            if clicado:
                break
    if not clicado:
        # Última tentativa com force=True (clicar mesmo com overlay)
        log("tentando click com force=True como ultima opcao")
        for b in botoes:
            if await b.is_visible():
                try:
                    await b.click(force=True, timeout=5000)
                    clicado = True
                    log("force click OK")
                    break
                except Exception as e:
                    log(f"force click falhou: {str(e)[:100]}")
    if not clicado:
        log("DESISTI - botao Pesquisar nao foi clicado")
        return None

    log("pesquisar clicado, aguardando resultado")
    await asyncio.sleep(2)

    # Verificar se apareceu o modal "Processo do PJe" — se sim, pular
    if await _detectar_processo_pje(page):
        log("processo eh do PJe — sem acesso, pulando")
        raise ProcessoPjeError(f"Processo {numero_processo} esta no PJe")

    # Aguardar resposta (link do CNJ aparecer)
    prefix_short = prefix.split('-')[0]  # ex: 0156129
    link_pattern = f"text=/{prefix_short}.*{sufix}/"
    try:
        await inner.locator(link_pattern).first.wait_for(state="visible", timeout=15000)
        log(f"link do resultado encontrado: {link_pattern}")
    except Exception as e:
        log(f"link nao encontrado: {e}")
        # Verificar de novo se eh PJe (pode aparecer demorado)
        if await _detectar_processo_pje(page):
            raise ProcessoPjeError(f"Processo {numero_processo} esta no PJe")
        return None

    # Clicar no link do resultado (navegação no mesmo SPA, mesma aba)
    try:
        await inner.locator(link_pattern).first.click(timeout=10000)
        log("clicou no link do resultado")
    except Exception as e:
        log(f"erro ao clicar link: {e}")
        # Pode ter modal interceptando
        await _clicar_prolongar_sessao(context)
        try:
            await inner.locator(link_pattern).first.click(timeout=10000)
            log("clicou no link (retry)")
        except Exception as e2:
            log(f"falhou click novamente: {e2}")
            return None
    await asyncio.sleep(3)

    # Refresh inner reference (URL pode ter mudado)
    inner = await _achar_frame_consulta(page)
    if inner is None:
        log("inner sumiu apos clicar link")
        return None
    log(f"inner apos clicar link: {inner.url[:80]}")

    # Clicar Visualizador (abre NOVA ABA)
    try:
        botao_visu = inner.locator(seletores.BOTAO_VISUALIZADOR).first
        await botao_visu.wait_for(state="visible", timeout=15000)
        log("botao Visualizador visivel")
    except Exception as e:
        log(f"botao Visualizador nao encontrado: {e}")
        return None

    # Garantir que não há overlay
    await _clicar_prolongar_sessao(context)

    # Clicar Visualizador. Pode acontecer:
    # (A) Sessão vinculada: abre NOVA ABA do visualizador imediatamente.
    # (B) Não vinculada: aparece modal de Solicitação de Acesso na MESMA aba.
    #     Preenchemos Motivo e clicamos botão, daí abre nova aba.
    visualizador = None
    try:
        async with context.expect_page(timeout=8000) as page_info:
            # Usar JS click pra evitar interceptação
            await botao_visu.evaluate("(el) => el.click()")
            log("click do Visualizador disparado (via JS)")
        visualizador = await page_info.value
        log(f"nova aba detectada: {visualizador.url[:80]}")
        await visualizador.wait_for_load_state("domcontentloaded", timeout=30000)
    except Exception as e:
        log(f"primeira tentativa nao abriu nova aba: {str(e)[:100]}")
        # Provavelmente apareceu modal Solicitação de Acesso
        await asyncio.sleep(2)
        # Verificar se modal apareceu
        modal_motivo = page.locator(seletores.MODAL_ACESSO_MOTIVO)
        for fr in page.frames:
            try:
                if await fr.locator(seletores.MODAL_ACESSO_MOTIVO).count() > 0:
                    if await fr.locator(seletores.MODAL_ACESSO_MOTIVO).first.is_visible():
                        log("modal Solicitacao de Acesso detectado")
                        break
            except: pass
        try:
            async with context.expect_page(timeout=15000) as page_info:
                ok = await _solicitar_acesso_se_aparecer(page)
                log(f"_solicitar_acesso_se_aparecer retornou: {ok}")
                if not ok:
                    return None
            visualizador = await page_info.value
            log(f"nova aba apos modal: {visualizador.url[:80]}")
            await visualizador.wait_for_load_state("domcontentloaded", timeout=30000)
        except Exception as e2:
            log(f"falha tambem no modal: {str(e2)[:100]}")
            return None

    # Confirmar que abriu visualizador (URL contém visproc)
    if seletores.VISUALIZADOR_URL_FRAGMENT not in visualizador.url:
        # Não é o visualizador esperado
        await visualizador.close()
        return None

    await asyncio.sleep(2)  # dar tempo do Angular carregar a árvore
    return visualizador


async def _test_consulta(numero):
    """Teste manual: navega até o visualizador de um processo específico."""
    pw = None
    browser = None
    try:
        pw, browser, context = await conectar_chrome_cdp()
        print(f"Consultando processo {numero}...")
        visu = await consultar_processo(context, numero, debug=True)
        if visu is None:
            print("FALHOU - visualizador nao abriu")
        else:
            print(f"SUCESSO - visualizador aberto: {visu.url[:120]}")
            print(f"Title: {await visu.title()}")
    finally:
        if browser:
            await browser.close()
        if pw:
            await pw.stop()


async def _expandir_arvore_completa(visualizador, max_iter=10):
    """Expande todos os mat-nested-tree-node colapsados até estabilizar.

    O visualizador do TJRJ tem nodes colapsados por padrão — sem expandir,
    os filhos não estão no DOM e :has-text não os encontra.
    """
    for _ in range(max_iter):
        mudou = await visualizador.evaluate("""
            () => {
                const nos = document.querySelectorAll(
                    'mat-nested-tree-node[aria-expanded="false"]'
                );
                if (nos.length === 0) return false;
                let count = 0;
                nos.forEach(n => {
                    // Procurar botão de expansão (mat-icon-button ou link com icone)
                    const btn = n.querySelector(
                        'button[aria-label*="xpand"], button.mat-icon-button, ' +
                        'mat-icon, [class*="expand"]'
                    );
                    if (btn) { btn.click(); count++; }
                });
                return count > 0;
            }
        """)
        if not mudou:
            break
        await asyncio.sleep(1.5)


def _ordenar_candidatos(candidatos_info):
    """Achata os matches da árvore numa lista ordenada e sem duplicatas de índices.

    Cada match traz `ids` = [idx_do_proprio_no, ...idx_dos_filhos]. O documento real
    pode ser o próprio nó OU um filho (quando o nó que casa é a capa de uma juntada).
    Preserva a ordem de descoberta e remove índices repetidos.
    """
    ordem = []
    vistos = set()
    for info in candidatos_info:
        for i in info.get("ids", []):
            if i not in vistos:
                vistos.add(i)
                ordem.append(i)
    return ordem


def _indices_fallback(ordem, ja_tentados, limite):
    """Filtra a ordem do fallback: remove índices já tentados no passo por rótulo,
    preserva a ordem de descoberta e corta no `limite`. Função pura (testável)."""
    out = []
    for i in ordem:
        if i in ja_tentados:
            continue
        out.append(i)
        if len(out) >= limite:
            break
    return out


def _faltam_vinculos(requisitorios, vinculos):
    """True se algum requisitório COM OFREQ não tem vínculo (OFREQ->precatório).

    Sem o vínculo, o requisitório vira 'SEM_VINCULO' (não liga ao nº do precatório).
    Os ofícios DEPJU de vínculo costumam estar como 'Petição' genérica no FIM da árvore,
    então o passo 1 (rótulo) acha o requisitório mas não o vínculo. Esta condição dispara
    o fallback do-fim-pra-frente pra ir buscar os vínculos que faltam. Função pura."""
    return any(req.get("ofreq") and req["ofreq"] not in vinculos for req in requisitorios)


def _parece_escaneado(texto, minimo=40):
    """True se o PDF não tem texto extraível suficiente (provável imagem/scan).

    PDFs escaneados fazem o pypdf devolver vazio ou só quebras de linha — nesse caso
    não dá pra extrair beneficiário/advogado e a peça vai pra revisão manual.
    """
    if not texto:
        return True
    return len(re.sub(r"\s", "", texto)) < minimo


# JS que casa o RÓTULO de cada nó da árvore (textContent) contra uma lista de padrões.
# Para cada match retorna o índice do nó + os índices dos nós-filhos (o documento real
# pode ser um filho — capa de juntada -> documento).
_JS_BUSCAR_CANDIDATOS = """
    (padroes) => {
        const nos = Array.from(document.querySelectorAll('mat-nested-tree-node'));
        const idxOf = new Map(nos.map((n, i) => [n, i]));
        const matches = [];
        nos.forEach((n, idx) => {
            const txt = (n.textContent || '').trim();
            const level = parseInt(n.getAttribute('aria-level') || '1');
            if (level < 2) return;
            for (const p of padroes) {
                if (txt.includes(p)) {
                    const ids = [idx];
                    n.querySelectorAll('mat-nested-tree-node').forEach(c => {
                        const ci = idxOf.get(c);
                        if (ci !== undefined) ids.push(ci);
                    });
                    matches.push({idx, level, padrao: p, ids});
                    break;
                }
            }
        });
        return matches;
    }
"""


async def _buscar_indices_candidatos(visualizador, padroes):
    """Casa os rótulos da árvore contra `padroes` e devolve a lista ordenada de
    índices (próprios + filhos), sem duplicatas. Reutilizado pelos dois passos
    (rótulo e fallback genérico)."""
    candidatos_info = await visualizador.evaluate(_JS_BUSCAR_CANDIDATOS, padroes)
    return _ordenar_candidatos(candidatos_info)


# JS do FALLBACK: casa apenas FOLHAS (nós sem filho-de-árvore = documentos reais),
# evitando os nós-pai de juntada que baixam o MESMO PDF do filho (duplicata que
# desperdiça o teto). Retorna índices em ordem DOM (do topo p/ baixo).
_JS_BUSCAR_LEAVES = """
    (padroes) => {
        const nos = Array.from(document.querySelectorAll('mat-nested-tree-node'));
        const out = [];
        nos.forEach((n, idx) => {
            if (n.querySelector('mat-nested-tree-node')) return;  // não é folha (é juntada)
            const level = parseInt(n.getAttribute('aria-level') || '1');
            if (level < 2) return;
            const txt = (n.textContent || '').trim();
            for (const p of padroes) {
                if (txt.includes(p)) { out.push(idx); break; }
            }
        });
        return out;
    }
"""


async def _buscar_indices_leaves(visualizador, padroes):
    """Como _buscar_indices_candidatos, mas só FOLHAS (documentos), sem os nós-pai
    de juntada. Usado pelo fallback para não gastar o teto baixando PDFs duplicados."""
    return await visualizador.evaluate(_JS_BUSCAR_LEAVES, padroes)


async def _baixar_e_classificar_um(visualizador, el, idx, pasta_temp,
                                   requisitorios, vinculos, ofreqs_vistos,
                                   log, numero_processo):
    """Baixa e classifica UM nó da árvore. Muta requisitorios/vinculos/ofreqs_vistos.

    Retorna True se o documento é RELEVANTE (requisitório ou vínculo pelo conteúdo —
    mesmo que duplicado/escaneado-falho), o que indica que ainda estamos no bloco de
    documentos do precatório. Retorna False para não-baixou / scan / 'não é nenhum dos
    dois'. Esse sinal alimenta o early-stop do loop.
    """
    try:
        try:
            rotulo = (await el.inner_text())[:70].replace("\n", " ").strip()
        except Exception:
            rotulo = "?"
        log(f"candidato idx={idx} rotulo={rotulo!r}: clicando")
        await el.evaluate("(e) => e.scrollIntoView({block: 'center'})")
        await asyncio.sleep(0.5)
        await el.click(force=True, timeout=10000)
        await asyncio.sleep(4)
    except Exception as e:
        log(f"  click falhou: {str(e)[:120]}")
        return False

    botao_salvar = visualizador.locator(seletores.BOTAO_SALVAR_COPIA).first
    try:
        await botao_salvar.wait_for(state="visible", timeout=5000)
    except Exception:
        log(f"  botao Salvar Copia nao apareceu — pulando")
        return False

    try:
        async with visualizador.expect_download(timeout=30000) as dl_info:
            await botao_salvar.evaluate("(el) => el.click()")
        download = await dl_info.value
        tmp_path = pasta_temp / f"_cand_{idx}_{download.suggested_filename}"
        await download.save_as(tmp_path)
        pdf_bytes = tmp_path.read_bytes()
        texto = _extrair_texto_pdf(pdf_bytes)
        log(f"  baixado: {tmp_path.stat().st_size} bytes")

        # PDF sem texto extraível = provável scan -> revisão manual
        if _parece_escaneado(texto):
            pasta_manual = pasta_temp.parent / "manual_revisar"
            pasta_manual.mkdir(parents=True, exist_ok=True)
            prefixo = f"{numero_processo}_" if numero_processo else ""
            tmp_path.replace(pasta_manual / f"{prefixo}{tmp_path.name}")
            log(f"  PDF sem texto (provável scan) — manual_revisar")
            return False

        # Documento de vínculo DEPRE/DEPJU: OFREQ -> precatório
        if eh_documento_vinculo(texto):
            v = extrair_vinculo_ofreq_precatorio(pdf_bytes)
            if v:
                vinculos[v[0]] = v[1]
                log(f"  vínculo: OFREQ {v[0]} -> precatório {v[1]}")
            tmp_path.unlink(missing_ok=True)  # não é entregável
            return True

        # Não é requisitório -> descartar
        if not eh_documento_requisitorio(texto):
            log(f"  PDF não é requisitório nem vínculo — descartando")
            tmp_path.unlink(missing_ok=True)
            return False

        # Requisitório: dedup por OFREQ
        ofreq = extrair_numero_ofreq(pdf_bytes)
        if ofreq and ofreq in ofreqs_vistos:
            log(f"  requisitório OFREQ {ofreq} repetido — descartando")
            tmp_path.unlink(missing_ok=True)
            return True  # ainda é um requisitório (bloco do precatório)

        benef = extrair_beneficiario_completo(pdf_bytes)
        if benef is None:
            pasta_manual = pasta_temp.parent / "manual_revisar"
            pasta_manual.mkdir(parents=True, exist_ok=True)
            tmp_path.replace(pasta_manual / tmp_path.name)
            log(f"  beneficiário não extraído — manual_revisar")
            return True  # é requisitório, só não extraiu — segue no bloco
        adv = extrair_advogado(pdf_bytes)
        if ofreq:
            ofreqs_vistos.add(ofreq)
        requisitorios.append({
            "ofreq": ofreq,
            "beneficiario": benef,
            "advogado": adv,
            "pdf_path": tmp_path,
            "pdf_bytes": pdf_bytes,
        })
        log(f"  requisitório coletado: OFREQ {ofreq} benef={benef['nome']}")
        return True
    except Exception as e:
        log(f"  erro download/parse: {e}")
        return False


async def _baixar_e_classificar_indices(visualizador, indices, pasta_temp,
                                        requisitorios, vinculos, ofreqs_vistos,
                                        log, numero_processo, parar_apos_misses=None):
    """Itera os índices baixando+classificando cada um. Muta os acumuladores.

    early-stop: se `parar_apos_misses` for dado, para depois de N downloads
    consecutivos IRRELEVANTES *após já ter achado* o bloco de documentos do precatório
    (no fallback do-fim-pra-frente, achados os requisitórios/vínculos, o resto da árvore
    são petições antigas inúteis). Misses ANTES do primeiro achado não contam.
    """
    todos_nos = visualizador.locator(seletores.ARVORE_ITEM_TREEITEM)
    achou_relevante = False
    misses_seguidos = 0
    for idx in indices:
        el = todos_nos.nth(idx)
        relevante = await _baixar_e_classificar_um(
            visualizador, el, idx, pasta_temp,
            requisitorios, vinculos, ofreqs_vistos, log, numero_processo)
        if relevante:
            achou_relevante = True
            misses_seguidos = 0
        elif achou_relevante:
            misses_seguidos += 1
            if parar_apos_misses is not None and misses_seguidos >= parar_apos_misses:
                log(f"  early-stop: {misses_seguidos} downloads irrelevantes após o bloco — parando")
                break


async def _baixar_e_extrair_pecas_ofreq(visualizador, pasta_temp, debug=False, numero_processo=None):
    """No visualizador, baixa as peças candidatas e as classifica:
    - requisitório (OFÍCIO REQUISITÓRIO) -> dados de beneficiário/advogado;
    - vínculo DEPRE/DEPJU ('gerou o precatório') -> mapa OFREQ -> precatório.

    Dois passos:
      1. RÓTULO — candidatos pelos padrões no rótulo da árvore (rápido, cobre a maioria).
      2. FALLBACK por CONTEÚDO — só se o passo 1 não rendeu nenhum requisitório: varre
         documentos genéricos ("Petição"/"Documento") e deixa o conteúdo classificar.
         Cobre cartórios que arquivam o requisitório como "Petição" (ex: 0110128-60.2015).

    Retorna (requisitorios, vinculos).
    """
    def log(msg):
        if debug:
            print(f"  [pecas] {msg}")
    pasta_temp = Path(pasta_temp)
    pasta_temp.mkdir(parents=True, exist_ok=True)
    requisitorios = []
    vinculos = {}
    ofreqs_vistos = set()

    # Expandir árvore inteira primeiro (nodes colapsados não estão no DOM)
    log("expandindo arvore...")
    await _expandir_arvore_completa(visualizador)
    await asyncio.sleep(2)

    # Passo 1: candidatos pelo RÓTULO da árvore — lidos do FIM pra o início (os
    # documentos do precatório, requisitório E vínculo, ficam nos eventos MAIS RECENTES;
    # ler do começo gasta o teto em documentos antigos e perde os do fim).
    ordem = await _buscar_indices_candidatos(visualizador, seletores.PADROES_PECA_REQUISITORIO)
    indices1 = ordem[::-1][: seletores.LIMITE_CANDIDATOS]
    log(f"passo 1 (rótulo, do fim): {len(indices1)} alvos")
    await _baixar_e_classificar_indices(
        visualizador, indices1, pasta_temp,
        requisitorios, vinculos, ofreqs_vistos, log, numero_processo)

    # Passo 2 (FALLBACK por conteúdo): dispara se o rótulo não achou requisitório OU se
    # achou requisitórios mas falta o vínculo de algum (os ofícios DEPJU costumam estar
    # como "Petição" genérica no FIM da árvore — senão o requisitório vira SEM_VINCULO).
    if not requisitorios or _faltam_vinculos(requisitorios, vinculos):
        log("passo 2 (fallback): requisitório sem rótulo ou vínculo faltando — varrendo do fim")
        # Só FOLHAS (documentos reais), sem os nós-pai de juntada — senão metade dos
        # downloads é o MESMO PDF (duplicata) e o teto se esgota antes de alcançar o alvo.
        ordem_g = await _buscar_indices_leaves(visualizador, seletores.PADROES_DOC_GENERICO)
        # Varrer do FIM da árvore pra frente: os requisitórios e ofícios DEPJU dos
        # precatórios ficam entre os eventos MAIS RECENTES (informação do usuário,
        # confirmada no 0110128 — eventos ~661-684 no fim). Assim alcançamos os alvos
        # nos primeiros downloads, dentro de um teto pequeno.
        ordem_g = ordem_g[::-1]
        indices2 = _indices_fallback(ordem_g, set(indices1), seletores.LIMITE_FALLBACK_GENERICO)
        log(f"passo 2 (fallback): {len(indices2)} documentos genéricos a inspecionar")
        await _baixar_e_classificar_indices(
            visualizador, indices2, pasta_temp,
            requisitorios, vinculos, ofreqs_vistos, log, numero_processo,
            parar_apos_misses=seletores.FALLBACK_EARLY_STOP_MISSES)

    log(f"total: {len(requisitorios)} requisitórios, {len(vinculos)} vínculos")
    return requisitorios, vinculos


def _eh_segunda_instancia(numero_processo):
    """True se o CNJ é de 2º grau (órgão de origem .0000).

    Processos de 2ª instância vivem no sistema legado ejud (não no SPA
    consultaprocessual) — são pulados e bucketados como 'legado_2inst' para
    uma etapa de extração separada.
    """
    return bool(numero_processo) and numero_processo.strip().endswith(".0000")


def _eh_falha_navegacao(status):
    """True para status que, em sequência, indicam navegador/sessão quebrado.

    Usado pelo circuit breaker. Resultados legítimos (ok, sem_requisitorio,
    legado_2inst, processo_pje) NÃO contam.
    """
    return status in ("erro_navegacao", "processo_nao_encontrado")


def casar_requisitorios_com_vinculos(requisitorios, vinculos, precatorios_do_processo):
    """Casa cada requisitório ao seu precatório pelo OFREQ.

    Retorna {precatorio: {beneficiario_nome, beneficiario_doc, advogado_nome,
    advogado_cpf, advogado_oab, status}}. Precatórios da lista que não receberam
    nenhum requisitório vinculado ficam com {"status": "REVISAR"}.
    """
    dados = {}
    for req in requisitorios:
        precatorio = vinculos.get(req["ofreq"])
        if not precatorio:
            continue
        benef = req["beneficiario"]
        adv = req.get("advogado")
        dados[precatorio] = {
            "beneficiario_nome": benef["nome"],
            "beneficiario_doc": formatar_doc(benef["doc"], benef["tipo_doc"]),
            "advogado_nome": adv["nome"] if adv else None,
            "advogado_cpf": formatar_doc(adv["cpf"], "CPF") if adv else None,
            "advogado_oab": adv["oab"] if adv else None,
            "status": "OK",
        }
    for precatorio in precatorios_do_processo:
        if precatorio not in dados:
            dados[precatorio] = {"status": "REVISAR"}
    return dados


async def processar_processo(context, precatorios_do_processo, numero_processo,
                              pasta_saida, pasta_tmp, debug=False):
    """Pipeline completo para um processo:
       consulta -> baixa OFREQ -> extrai dados -> renomeia PDFs.

    Retorna dict:
      - status: "ok" | "sem_requisitorio" | "processo_nao_encontrado" | "erro_*"
      - arquivos: list[str] (nomes dos PDFs)
      - dados: dict {precatorio: {beneficiario_nome, beneficiario_doc, advogado_nome,
                                   advogado_cpf, advogado_oab}}
    """
    # P5b: processos de 2ª instância (.0000) vivem no ejud legado — pular antes de
    # navegar evita que prendam o iframe e bucketar para extração separada.
    if _eh_segunda_instancia(numero_processo):
        return {"status": "legado_2inst", "arquivos": [], "dados": {},
                "motivo": "2a instancia (ejud) - extracao separada"}
    visu = None
    try:
        try:
            visu = await consultar_processo(context, numero_processo, debug=debug)
        except ProcessoPjeError:
            return {"status": "processo_pje", "arquivos": [], "dados": {}, "motivo": "no sistema PJe"}
        if visu is None:
            return {"status": "processo_nao_encontrado", "arquivos": [], "dados": {}}

        # CRÍTICO: trazer a aba pra frente para que o Angular renderize a árvore.
        # Sem isso, a aba fica em background e a árvore vem vazia.
        await visu.bring_to_front()
        await asyncio.sleep(2)

        # Aguardar a árvore de peças carregar (com bring_to_front, vai renderizar)
        try:
            await visu.locator(seletores.ARVORE_ITEM_TREEITEM).first.wait_for(
                state="visible", timeout=30000
            )
            # Aguardar mais tempo para a árvore inteira renderizar (Angular renderiza em batches)
            await asyncio.sleep(5)
        except Exception:
            return {"status": "sem_requisitorio", "arquivos": [], "dados": {}, "motivo": "arvore_nao_carregou"}

        # Esperar a árvore "estabilizar" — checar contagem várias vezes até parar de crescer
        contagem_anterior = 0
        for _ in range(8):
            n_atual = await visu.locator(seletores.ARVORE_ITEM_TREEITEM).count()
            if n_atual == contagem_anterior and n_atual > 0:
                break
            contagem_anterior = n_atual
            await asyncio.sleep(1)

        requisitorios, vinculos = await _baixar_e_extrair_pecas_ofreq(
            visu, pasta_tmp, debug=debug, numero_processo=numero_processo)
        if not requisitorios:
            return {"status": "sem_requisitorio", "arquivos": [], "dados": {}}

        pasta_saida = Path(pasta_saida)
        pasta_saida.mkdir(parents=True, exist_ok=True)

        # Nomear cada PDF pelo precatório VERDADEIRO (via vínculo OFREQ)
        arquivos_finais = []
        for req in requisitorios:
            precatorio = vinculos.get(req["ofreq"]) or "SEM_VINCULO"
            destino = gerar_nome_arquivo(
                precatorio, req["beneficiario"]["nome"], pasta_saida)
            req["pdf_path"].replace(destino)
            arquivos_finais.append(destino.name)

        # Join OFREQ -> precatório, montando os dados por precatório
        dados = casar_requisitorios_com_vinculos(
            requisitorios, vinculos, precatorios_do_processo)

        return {"status": "ok", "arquivos": arquivos_finais, "dados": dados}

    except LoginExpiradoError:
        raise
    except Exception as e:
        return {
            "status": "erro_navegacao",
            "arquivos": [], "dados": {},
            "motivo": f"{type(e).__name__}: {e}",
        }
    finally:
        if visu is not None and not visu.is_closed():
            try:
                await visu.close()
            except Exception:
                pass


async def _test_processar(numero):
    """Pipeline completo de um processo de teste (lança Edge via pipe, igual ao main)."""
    pw = None
    context = None
    try:
        pw = await async_playwright().start()
        context = await abrir_context(pw, headless=False)
        print("Edge lançado via pipe (sem porta de debug — AV-friendly)")
        if not await verificar_sessao(context):
            if not await aguardar_login_no_browser(context):
                print("Login não confirmado. Saindo.", file=sys.stderr)
                return
        print("Sessão confirmada.")
        pasta_saida = Path.home() / "Downloads" / "Precatórios_Requisitórios"
        pasta_saida.mkdir(parents=True, exist_ok=True)
        pasta_tmp = PROJETO_DIR / "_tmp_downloads"
        print(f"Processando {numero}...")
        resultado = await processar_processo(
            context, precatorios_do_processo=["TESTE"],
            numero_processo=numero, pasta_saida=pasta_saida, pasta_tmp=pasta_tmp,
            debug=True,
        )
        print(f"Resultado: {resultado}")
    finally:
        if context:
            try:
                await context.close()
            except Exception:
                pass
        if pw:
            await pw.stop()


# =============================================================================
# Orquestração (Task 10): Watchdog do modal de sessão + worker pool + CLI
# =============================================================================

import argparse
import csv
from datetime import datetime


PASTA_DOWNLOADS_HOME = Path.home() / "Downloads"
SAIDA_PDFS_PADRAO = PASTA_DOWNLOADS_HOME / "Precatórios_Requisitórios"
ENTRADA_PADRAO = PASTA_DOWNLOADS_HOME / "Precatórios 2027 - Atualizado.xlsx"

# Circuit breaker: nº de falhas de navegação SEGUIDAS que aborta o run (sessão caiu)
LIMITE_FALHAS_SEGUIDAS = 10

# Deadline por processo. Um processo saudável leva 30-90s (runbook); multi-precatório
# com vários PDFs pode chegar a ~3min. Acima disso = página travada (ex: variante de
# PJe que prende um await do Playwright SEM timeout). Sem este deadline, um único
# processo travado congela o lote INTEIRO (CPU ~0, sem progresso). Ao estourar vira
# erro_navegacao → conta pro circuit breaker e é retomável via --apenas-erros.
PROCESSO_TIMEOUT_SEG = 240


def caminho_saida_padrao(entrada):
    return entrada.with_name(entrada.stem + " - Etapa2.xlsx")


async def _processar_com_timeout(context, precatorios_do_processo, numero_processo,
                                 pasta_saida, pasta_tmp,
                                 timeout_seg=PROCESSO_TIMEOUT_SEG):
    """Envolve processar_processo num deadline para que nenhum processo isolado
    trave o lote. Na maioria das vezes os awaits internos têm timeout próprio, mas
    a COMPOSIÇÃO não — basta um await sem timeout numa página wedged pra congelar
    tudo. Aqui o asyncio.wait_for cancela o processo travado e devolve erro_navegacao
    (o loop principal segue pro próximo). LoginExpiradoError continua propagando
    (wait_for re-levanta exceções que não são TimeoutError) para o main parar e salvar.
    """
    try:
        return await asyncio.wait_for(
            processar_processo(context, precatorios_do_processo, numero_processo,
                               pasta_saida, pasta_tmp),
            timeout=timeout_seg,
        )
    except asyncio.TimeoutError:
        return {
            "status": "erro_navegacao",
            "arquivos": [], "dados": {},
            "motivo": f"timeout: processo excedeu {timeout_seg}s (travado)",
        }


def parse_args(argv=None):
    p = argparse.ArgumentParser(
        description="Pipeline TJRJ: baixa requisitórios + extrai dados + atualiza planilha."
    )
    p.add_argument("--entrada", type=Path, default=ENTRADA_PADRAO)
    p.add_argument("--saida", type=Path, default=None)
    p.add_argument("--pasta-pdfs", type=Path, default=SAIDA_PDFS_PADRAO)
    p.add_argument("--saldo-minimo", type=float, default=200000.0)
    p.add_argument("--workers", type=int, default=1,
                   help="Abas paralelas. CUIDADO: o TJRJ pode bloquear se >1")
    p.add_argument("--cdp-port", type=int, default=CDP_PORTA_PADRAO)
    p.add_argument("--limit", type=int, default=None)
    p.add_argument("--reset", action="store_true")
    p.add_argument("--apenas-erros", action="store_true")
    p.add_argument("--processo", type=str, default=None)
    return p.parse_args(argv)


def precisa_processar(processo, resultados, apenas_erros):
    if processo not in resultados:
        return True
    status = resultados[processo].get("status", "")
    if apenas_erros and status != "ok":
        return True
    # Run normal: NÃO refaz nada que já está no checkpoint (inclusive erro_navegacao).
    # Os erro_navegacao costumam ser FALSOS erros (a sessão TJRJ caiu, não o processo);
    # retomar inline a cada resume só re-lê os mesmos processos. Ficam para um passe
    # final único com --apenas-erros, depois que o frontier terminar.
    return False


async def _watchdog_sessao(context, stop_event):
    """Task assíncrona: a cada N segundos, procura modal de sessão inativa em todas
    as abas TJRJ e clica 'Prolongar sessão'. Encerra quando stop_event é setado.
    """
    while not stop_event.is_set():
        try:
            n = await _clicar_prolongar_sessao(context)
            if n > 0:
                print(f"\n[watchdog] Prolongou sessão em {n} aba(s)")
        except Exception:
            pass
        # Esperar com cancelamento responsivo
        try:
            await asyncio.wait_for(stop_event.wait(), timeout=seletores.INTERVALO_POLLING_MODAL_SEG)
        except asyncio.TimeoutError:
            continue


def _print_progresso(concluidos, total, inicio):
    pct = concluidos / total * 100 if total else 0
    decorrido = (datetime.now() - inicio).total_seconds()
    taxa = concluidos / decorrido if decorrido > 0 else 0
    eta = (total - concluidos) / taxa if taxa > 0 else 0
    eta_str = f"{int(eta//60)}min{int(eta%60):02d}s"
    barra = "█" * int(30 * concluidos / total) + "░" * (30 - int(30 * concluidos / total))
    print(f"\r  [{barra}] {concluidos}/{total} ({pct:.1f}%)  ETA {eta_str}", end="", flush=True)


async def main():
    args = parse_args()
    if not args.entrada.exists():
        print(f"ERRO: arquivo não encontrado: {args.entrada}", file=sys.stderr)
        return 1

    args.pasta_pdfs.mkdir(parents=True, exist_ok=True)
    SCREENSHOTS_DIR.mkdir(parents=True, exist_ok=True)
    saida_xlsx = args.saida or caminho_saida_padrao(args.entrada)

    if args.reset and CHECKPOINT_PATH.exists():
        CHECKPOINT_PATH.unlink()
        print(f"Checkpoint apagado: {CHECKPOINT_PATH}")

    resultados = carregar_checkpoint_dl(CHECKPOINT_PATH)
    print(f"Checkpoint: {len(resultados)} processos")

    print(f"Lendo {args.entrada}...")
    pares = filtrar_precatorios(args.entrada, args.saldo_minimo)
    print(f"  {len(pares)} precatórios com saldo >= R$ {args.saldo_minimo:,.2f}")

    # Agrupar precatórios por processo
    mapa_proc_to_precs = {}
    for prec, proc in pares:
        mapa_proc_to_precs.setdefault(proc, []).append(prec)
    print(f"  {len(mapa_proc_to_precs)} processos únicos")

    if args.processo:
        if args.processo in mapa_proc_to_precs:
            mapa_proc_to_precs = {args.processo: mapa_proc_to_precs[args.processo]}
        else:
            mapa_proc_to_precs = {args.processo: ["DEBUG"]}

    pendentes = [
        (proc, precs) for proc, precs in mapa_proc_to_precs.items()
        if precisa_processar(proc, resultados, args.apenas_erros)
    ]
    if args.limit:
        pendentes = pendentes[: args.limit]
    print(f"  {len(pendentes)} pendentes para processar")

    if not pendentes:
        print("Nada a fazer.")
        return 0

    # Reconstruir dados acumulados do checkpoint
    dados_acumulados = {}
    for proc, r in resultados.items():
        if r.get("status") == "ok" and proc in mapa_proc_to_precs:
            dados_acumulados.update(r.get("dados", {}))
    print(f"  {len(dados_acumulados)} precatórios com dados extraídos no checkpoint")

    inicio = datetime.now()
    pw = None
    stop_event = asyncio.Event()
    watchdog = None

    context = None
    try:
        pw = await async_playwright().start()
        context = await abrir_context(pw, headless=False)
        print("Edge lançado via pipe (sem porta de debug — AV-friendly)")
        # Login interativo se necessário; o Edge fica ABERTO durante todo o run
        # (mantém a sessão viva — cookies de sessão não sobrevivem a fechar o browser).
        if not await verificar_sessao(context):
            if not await aguardar_login_no_browser(context):
                print("Login não confirmado. Saindo.", file=sys.stderr)
                return 4
        print("Sessão confirmada.")
        watchdog = asyncio.create_task(_watchdog_sessao(context, stop_event))

        # Processar sequencialmente (workers=1 default — pra evitar bloqueio)
        # Para múltiplos workers, cada um precisaria de sua aba dedicada
        falhas_seguidas = 0  # circuit breaker (P2)
        for idx, (numero_processo, precatorios) in enumerate(pendentes, 1):
            try:
                pasta_tmp = PROJETO_DIR / "_tmp_downloads"
                resultado = await _processar_com_timeout(
                    context, precatorios, numero_processo,
                    args.pasta_pdfs, pasta_tmp,
                )
                resultados[numero_processo] = {
                    "status": resultado["status"],
                    "arquivos": resultado.get("arquivos", []),
                    "motivo": resultado.get("motivo"),
                    "dados": resultado.get("dados", {}),
                }
                if resultado.get("dados"):
                    dados_acumulados.update(resultado["dados"])

                if idx % 10 == 0 or idx == len(pendentes):
                    salvar_checkpoint_dl(CHECKPOINT_PATH, resultados)
                    atualizar_planilha(args.entrada, saida_xlsx, dados_acumulados)
                    _print_progresso(idx, len(pendentes), inicio)
            except LoginExpiradoError:
                print(f"\nSessão expirou no processo {numero_processo}. Salvando e saindo.")
                salvar_checkpoint_dl(CHECKPOINT_PATH, resultados)
                atualizar_planilha(args.entrada, saida_xlsx, dados_acumulados)
                return 2
            except Exception as e:
                resultados[numero_processo] = {
                    "status": "erro_navegacao", "arquivos": [], "dados": {},
                    "motivo": f"{type(e).__name__}: {e}",
                }
                print(f"\n[ERRO] {numero_processo}: {e}")

            # Circuit breaker (P2): muitas falhas SEGUIDAS = navegador/sessão caiu.
            # Para limpo em vez de moer centenas de processos como erro.
            if _eh_falha_navegacao(resultados[numero_processo]["status"]):
                falhas_seguidas += 1
            else:
                falhas_seguidas = 0
            if falhas_seguidas >= LIMITE_FALHAS_SEGUIDAS:
                print(f"\n[CIRCUIT BREAKER] {falhas_seguidas} falhas de navegação seguidas — "
                      f"o navegador/sessão provavelmente caiu. Salvando e parando.\n"
                      f"  Relogue no Edge (Portal de Serviços, perfil Advogado) e rode de novo "
                      f"— retoma de onde parou.")
                salvar_checkpoint_dl(CHECKPOINT_PATH, resultados)
                atualizar_planilha(args.entrada, saida_xlsx, dados_acumulados)
                return 3

        # Salvamento final
        salvar_checkpoint_dl(CHECKPOINT_PATH, resultados)
        atualizar_planilha(args.entrada, saida_xlsx, dados_acumulados)
    finally:
        stop_event.set()
        if watchdog:
            try:
                await asyncio.wait_for(watchdog, timeout=5)
            except Exception:
                watchdog.cancel()
        if context:
            try:
                await context.close()
            except Exception:
                pass
        if pw:
            await pw.stop()

    print()
    _print_relatorio(mapa_proc_to_precs, resultados, saida_xlsx, args.pasta_pdfs, inicio)
    return 0


def _print_relatorio(mapa, resultados, saida_xlsx, pasta_pdfs, inicio):
    ok = sem_req = nao_enc = pje = erro_rede = erro_nav = erro_pars = legado = 0
    erros_detalhe = []
    total_pdfs = 0
    for proc in mapa:
        r = resultados.get(proc, {})
        status = r.get("status", "ausente")
        if status == "ok":
            ok += 1
            total_pdfs += len(r.get("arquivos", []))
        elif status == "sem_requisitorio":
            sem_req += 1
        elif status == "legado_2inst":
            legado += 1
        elif status == "processo_nao_encontrado":
            nao_enc += 1
        elif status == "processo_pje":
            pje += 1
        elif status == "erro_rede":
            erro_rede += 1
            erros_detalhe.append((proc, status, r.get("motivo", "")))
        elif status == "erro_navegacao":
            erro_nav += 1
            erros_detalhe.append((proc, status, r.get("motivo", "")))
        elif status == "erro_parsing":
            erro_pars += 1
            erros_detalhe.append((proc, status, r.get("motivo", "")))

    decorrido = (datetime.now() - inicio).total_seconds()
    print()
    print("=" * 60)
    print(f" CONCLUÍDO  ({int(decorrido//60)}min{int(decorrido%60)}s)")
    print(f" Processos OK              : {ok}  ({total_pdfs} PDFs)")
    print(f" Sem requisitório          : {sem_req}")
    print(f" 2ª instância (ejud/skip)  : {legado}")
    print(f" Processo não encontrado   : {nao_enc}")
    print(f" Processos no PJe (skip)   : {pje}")
    print(f" Erros de rede             : {erro_rede}")
    print(f" Erros de navegação        : {erro_nav}")
    print(f" Erros de parsing          : {erro_pars}")
    print(f" Planilha de saída         : {saida_xlsx}")
    print(f" Pasta de PDFs             : {pasta_pdfs}")
    print("=" * 60)
    if erros_detalhe:
        with open(ERROS_CSV, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["Processo", "Status", "Motivo"])
            w.writerows(erros_detalhe)
        print(f" Log de erros: {ERROS_CSV}")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] in ("--login", "--test-login"):
        asyncio.run(_test_login_manual())
    elif len(sys.argv) > 1 and sys.argv[1] == "--test-cdp":
        asyncio.run(_test_cdp_manual())
    elif len(sys.argv) > 2 and sys.argv[1] == "--test-consulta":
        asyncio.run(_test_consulta(sys.argv[2]))
    elif len(sys.argv) > 2 and sys.argv[1] == "--test-processar":
        asyncio.run(_test_processar(sys.argv[2]))
    else:
        sys.exit(asyncio.run(main()))
        print("(main completo será implementado em task posterior)")
