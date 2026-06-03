"""Atualiza saldos de precatórios consultando a API do TJRJ."""
import os
# Remover SSLKEYLOGFILE se apontar para arquivo virtual inacessível (sandbox do Windows)
os.environ.pop("SSLKEYLOGFILE", None)

import re
import time
import json
from pathlib import Path
from openpyxl import load_workbook
import argparse
import csv
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import requests

REGEX_PRECATORIO = re.compile(r"^\d{4}\.\d+-\d+$")


def extrair_precatorios(caminho_xlsx):
    """Lê coluna B do xlsx e retorna lista de (linha_excel, numero).

    Filtra apenas linhas cuja coluna B casa com o padrão de número de precatório.
    Pula a linha 1 (cabeçalho).
    """
    wb = load_workbook(caminho_xlsx, read_only=True, data_only=True)
    ws = wb.active
    precatorios = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 2:
            continue
        numero = row[1]
        if isinstance(numero, str) and REGEX_PRECATORIO.match(numero):
            precatorios.append((row_idx, numero))
    wb.close()
    return precatorios


URL_API = "https://www3.tjrj.jus.br/PortalConhecimento/api/precatorios/consultaPosicao"


def consultar_saldo(numero, session, max_tentativas=3, timeout=30):
    """Consulta o saldo do precatório no TJRJ.

    Retorna float em caso de sucesso, ou string com código de erro:
      "NAO_ENCONTRADO" — HTTP 404
      "SEM_SALDO"      — HTTP 200 mas Saldo é null
      "RESPOSTA_INVALIDA" — JSON malformado
      "ERRO_REDE"      — timeout / 5xx após N tentativas
    """
    for tentativa in range(1, max_tentativas + 1):
        try:
            resp = session.get(URL_API, params={"numeroPrecatorio": numero}, timeout=timeout)
            if resp.status_code == 404:
                return "NAO_ENCONTRADO"
            if resp.status_code >= 500:
                raise IOError(f"HTTP {resp.status_code}")
            if resp.status_code != 200:
                return "RESPOSTA_INVALIDA"
            try:
                dados = resp.json()
            except ValueError:
                return "RESPOSTA_INVALIDA"
            saldo = dados.get("Saldo")
            if saldo is None:
                return "SEM_SALDO"
            return float(saldo)
        except (IOError, OSError):
            if tentativa < max_tentativas:
                time.sleep(2 ** (tentativa - 1))
            continue
    return "ERRO_REDE"


def carregar_checkpoint(caminho):
    """Carrega checkpoint JSON. Retorna dict vazio se arquivo não existir."""
    caminho = Path(caminho)
    if not caminho.exists():
        return {}
    with open(caminho, "r", encoding="utf-8") as f:
        return json.load(f)


def salvar_checkpoint(caminho, dados):
    """Salva checkpoint de forma atômica: escreve .tmp e faz rename."""
    caminho = Path(caminho)
    tmp = caminho.with_suffix(caminho.suffix + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    tmp.replace(caminho)


COLUNA_SALDO = 8  # coluna H
FORMATO_MOEDA = 'R$ #,##0.00'


def escrever_xlsx(entrada, saida, precatorios, resultados):
    """Carrega xlsx de entrada, preenche coluna H e salva em saída.

    O arquivo de entrada não é modificado.
    Resultados não-numéricos (strings de erro) são ignorados.
    """
    entrada = Path(entrada)
    saida = Path(saida)
    wb = load_workbook(entrada)
    ws = wb.active
    for linha, numero in precatorios:
        valor = resultados.get(numero)
        celula = ws.cell(row=linha, column=COLUNA_SALDO)
        celula.number_format = FORMATO_MOEDA
        if isinstance(valor, (int, float)):
            celula.value = float(valor)
    wb.save(saida)
    wb.close()


PASTA_DOWNLOADS = Path.home() / "Downloads"
ENTRADA_PADRAO = PASTA_DOWNLOADS / "Precatórios 2027.xlsx"
SAIDA_PADRAO_SUFIX = " - Atualizado.xlsx"
CHECKPOINT_PADRAO = Path(__file__).parent / "saldos_checkpoint.json"
ERROS_CSV = Path(__file__).parent / "erros_consulta.csv"


def _formatar_eta(segundos):
    if segundos < 60:
        return f"{int(segundos)}s"
    minutos, seg = divmod(int(segundos), 60)
    return f"{minutos}min{seg:02d}s"


def main(argv=None):
    parser = argparse.ArgumentParser(description="Atualiza saldos de precatórios via API do TJRJ.")
    parser.add_argument("--entrada", type=Path, default=ENTRADA_PADRAO)
    parser.add_argument("--saida", type=Path, default=None)
    parser.add_argument("--checkpoint", type=Path, default=CHECKPOINT_PADRAO)
    parser.add_argument("--workers", type=int, default=20)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--reset", action="store_true", help="Apaga checkpoint antes")
    parser.add_argument("--apenas-erros", action="store_true", help="Re-consulta entradas com erro")
    parser.add_argument("--so-aplicar", action="store_true", help="Pula consulta, só aplica checkpoint")
    args = parser.parse_args(argv)

    if not args.entrada.exists():
        print(f"ERRO: arquivo não encontrado: {args.entrada}", file=sys.stderr)
        return 1

    saida = args.saida or args.entrada.with_name(args.entrada.stem + SAIDA_PADRAO_SUFIX)

    if args.reset and args.checkpoint.exists():
        args.checkpoint.unlink()
        print(f"Checkpoint apagado: {args.checkpoint}")

    print(f"Lendo {args.entrada}...")
    precatorios = extrair_precatorios(args.entrada)
    print(f"Encontrados {len(precatorios)} precatórios.")

    resultados = carregar_checkpoint(args.checkpoint)
    print(f"Checkpoint atual: {len(resultados)} entradas.")

    if not args.so_aplicar:
        def precisa_consultar(numero):
            if numero not in resultados:
                return True
            valor = resultados[numero]
            if valor == "ERRO_REDE":
                return True
            if args.apenas_erros and isinstance(valor, str):
                return True
            return False

        pendentes = [(linha, n) for linha, n in precatorios if precisa_consultar(n)]
        if args.limit:
            pendentes = pendentes[: args.limit]

        if pendentes:
            print(f"Consultando {len(pendentes)} precatórios com {args.workers} workers...")
            _consultar_em_paralelo(pendentes, resultados, args.checkpoint, args.workers)
        else:
            print("Nada pendente para consultar.")

    print(f"Escrevendo {saida}...")
    escrever_xlsx(args.entrada, saida, precatorios, resultados)

    _imprimir_relatorio(precatorios, resultados, saida)
    return 0


def _consultar_em_paralelo(pendentes, resultados, caminho_checkpoint, workers):
    inicio = datetime.now()
    total = len(pendentes)
    concluidos = 0

    with requests.Session() as session:
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futuros = {pool.submit(consultar_saldo, n, session): (linha, n) for linha, n in pendentes}
            for futuro in as_completed(futuros):
                linha, numero = futuros[futuro]
                try:
                    resultados[numero] = futuro.result()
                except Exception as e:
                    resultados[numero] = f"ERRO_INESPERADO:{type(e).__name__}"

                concluidos += 1
                if concluidos % 100 == 0 or concluidos == total:
                    salvar_checkpoint(caminho_checkpoint, resultados)
                    _imprimir_progresso(concluidos, total, inicio)

    salvar_checkpoint(caminho_checkpoint, resultados)
    print()


def _imprimir_progresso(concluidos, total, inicio):
    decorrido = (datetime.now() - inicio).total_seconds()
    pct = concluidos / total * 100
    taxa = concluidos / decorrido if decorrido > 0 else 0
    eta = (total - concluidos) / taxa if taxa > 0 else 0
    barra_size = 30
    preenchido = int(barra_size * concluidos / total)
    barra = "█" * preenchido + "░" * (barra_size - preenchido)
    print(f"\r  [{barra}] {concluidos}/{total} ({pct:.1f}%) | {taxa:.1f}/s | ETA {_formatar_eta(eta)}",
          end="", flush=True)


def _imprimir_relatorio(precatorios, resultados, saida):
    sucesso = sem_saldo = nao_encontrado = erro_rede = invalido = outros = 0
    erros_detalhe = []
    for linha, numero in precatorios:
        valor = resultados.get(numero)
        if isinstance(valor, (int, float)):
            sucesso += 1
        elif valor == "SEM_SALDO":
            sem_saldo += 1
            erros_detalhe.append((linha, numero, "SEM_SALDO"))
        elif valor == "NAO_ENCONTRADO":
            nao_encontrado += 1
            erros_detalhe.append((linha, numero, "NAO_ENCONTRADO"))
        elif valor == "ERRO_REDE":
            erro_rede += 1
            erros_detalhe.append((linha, numero, "ERRO_REDE"))
        elif valor == "RESPOSTA_INVALIDA":
            invalido += 1
            erros_detalhe.append((linha, numero, "RESPOSTA_INVALIDA"))
        elif valor is None:
            outros += 1
        else:
            outros += 1
            erros_detalhe.append((linha, numero, str(valor)))

    print()
    print("=" * 50)
    print(f" CONCLUÍDO")
    print(f" Arquivo: {saida}")
    print(f" Sucesso          : {sucesso}")
    print(f" Sem saldo        : {sem_saldo}")
    print(f" Não encontrados  : {nao_encontrado}")
    print(f" Erros de rede    : {erro_rede}")
    print(f" Resposta inválida: {invalido}")
    if outros:
        print(f" Outros           : {outros}")
    print("=" * 50)

    if erros_detalhe:
        with open(ERROS_CSV, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["Linha", "Numero", "Motivo"])
            w.writerows(erros_detalhe)
        print(f" Log de erros: {ERROS_CSV}")


if __name__ == "__main__":
    sys.exit(main())
